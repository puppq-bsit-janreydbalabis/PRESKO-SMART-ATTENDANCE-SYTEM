from kivy.app import App
from kivy.clock import Clock
from kivy.core.image import Image as CoreImage
from kivy.core.window import Window
from kivy.graphics import (
    Color,
    RoundedRectangle,
    Rectangle,
    Line,
    Ellipse,
    StencilPush,
    StencilUse,
    StencilPop,
)
from kivy.uix.screenmanager import Screen, ScreenManager, SlideTransition
from kivy.uix.spinner import SpinnerOption
from kivy.properties import (
    ListProperty,
    StringProperty,
    BooleanProperty,
    ObjectProperty,
)
from kivy.uix.behaviors import ButtonBehavior
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.image import Image, AsyncImage
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.uix.spinner import Spinner
from kivy.uix.textinput import TextInput
from kivy.uix.widget import Widget
from kivy.core.window import Window
from functools import partial
from kivy.utils import platform, get_random_color
from kivy.metrics import dp
from kivy.uix.relativelayout import RelativeLayout
from kivy.graphics.texture import Texture

# --- NEW: KivyMD Imports ---
# The main app class must now inherit from MDApp
from kivymd.app import MDApp
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.button import MDIconButton
from kivymd.uix.label import MDIcon
from kivymd.uix.navigationdrawer import (
    MDNavigationDrawer,
    MDNavigationDrawerMenu,
    MDNavigationDrawerHeader,
    MDNavigationDrawerItem,
)


# Set window size for desktop testing to simulate mobile view
if platform not in ("android", "ios"):
    Window.size = (360, 640)  # Typical mobile resolution (e.g., 360x640)
    Window.fullscreen = False  # Disable fullscreen for desktop testing
else:
    Window.fullscreen = "auto"  # Keep fullscreen for actual mobile devices

from datetime import datetime, timedelta
from io import BytesIO
import calendar
import cv2
import os
import qrcode
import random
import string
import time
import sqlite3
import ctypes
from pyzbar import pyzbar
import numpy as np

# ### MODIFIED ###: Replaced csv with openpyxl for formatted Excel reports
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


# Load libzbar-64.dll explicitly for Windows (desktop testing)
if platform == "win":
    zbar_dll_path = os.path.join(os.path.dirname(__file__), "libzbar-64.dll")
    if os.path.exists(zbar_dll_path):
        try:
            ctypes.WinDLL(zbar_dll_path)
        except Exception as e:
            print(f"Warning: Could not load libzbar-64.dll. Error: {e}")
    else:
        print(
            f"Warning: {zbar_dll_path} not found. Ensure libzbar-64.dll is in the script directory."
        )


conn = sqlite3.connect("presko.db")
cursor = conn.cursor()

try:
    cursor.execute("ALTER TABLE subjects ADD COLUMN created_by TEXT")
except sqlite3.OperationalError:
    # This means the column already exists, ignore
    pass

# ### MODIFIED ###: Add the new late_cutoff_time column to the qr_validity table
try:
    cursor.execute("ALTER TABLE qr_validity ADD COLUMN late_cutoff_time TEXT")
except sqlite3.OperationalError:
    # Column already exists
    pass

conn.commit()
conn.close()


def init_database():
    conn = sqlite3.connect("presko.db")
    cursor = conn.cursor()

    # ### MODIFIED ###: Add late_cutoff_time to the qr_validity table definition.
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS qr_validity (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_code TEXT NOT NULL,
            date TEXT NOT NULL,
            manual_code TEXT NOT NULL,
            start_time TEXT NOT NULL,
            end_time TEXT NOT NULL,
            late_cutoff_time TEXT NOT NULL,
            UNIQUE(subject_code, date)
        )
    """
    )

    conn.commit()
    conn.close()


# Create or connect to the database
conn = sqlite3.connect("presko.db")
cursor = conn.cursor()

cursor.execute(
    """
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    username TEXT UNIQUE NOT NULL,
    password TEXT NOT NULL,
    role TEXT NOT NULL CHECK(role IN ('Student', 'Professor', 'Admin'))
)
"""
)

# Auto-create Admin account if not exists
cursor.execute(
    "SELECT * FROM users WHERE username = ? AND role = ?", ("admin", "Admin")
)
admin_exists = cursor.fetchone()

if not admin_exists:
    cursor.execute(
        """
        INSERT INTO users (name, username, password, role)
        VALUES (?, ?, ?, ?)
    """,
        ("Admin Account", "admin", "admin123", "Admin"),
    )
    conn.commit()

# SUBJECTS table
cursor.execute(
    """
CREATE TABLE IF NOT EXISTS subjects (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    subject_code TEXT NOT NULL UNIQUE,
    subject_name TEXT NOT NULL,
    professor_name TEXT NOT NULL,
    schedule TEXT NOT NULL,
    section TEXT NOT NULL,
    created_by TEXT NOT NULL
)
"""
)

# STUDENT_SUBJECTS table
cursor.execute(
    """
CREATE TABLE IF NOT EXISTS student_subjects (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    student_id TEXT NOT NULL,
    subject_code TEXT NOT NULL,
    UNIQUE(student_id, subject_code)
)
"""
)

# ### MODIFIED ###: Added the late_cutoff_time to ensure the table is correct on startup.
cursor.execute(
    """
CREATE TABLE IF NOT EXISTS qr_validity (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    subject_code TEXT NOT NULL,
    date TEXT NOT NULL,
    manual_code TEXT NOT NULL,
    start_time TEXT NOT NULL,
    end_time TEXT NOT NULL,
    late_cutoff_time TEXT NOT NULL,
    UNIQUE(subject_code, date)
)
"""
)

# ATTENDANCE table
cursor.execute(
    """
CREATE TABLE IF NOT EXISTS attendance (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    student_id TEXT NOT NULL,
    subject_code TEXT NOT NULL,
    date TEXT NOT NULL,
    status TEXT CHECK(status IN ('present', 'absent', 'late'))
)
"""
)

cursor.execute(
    """
CREATE TABLE IF NOT EXISTS feedback (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    sender_id TEXT NOT NULL,
    sender_role TEXT NOT NULL,
    message TEXT NOT NULL,
    timestamp TEXT NOT NULL
)
"""
)

conn.commit()
conn.close()


# ##########################################################################
# ###################### START: MODIFIED CODE BLOCK ########################
# ##########################################################################


# ### NEW HELPER CLASS: MenuDrawerButton ###
# This class creates the custom buttons for the navigation drawer,
# ensuring they look like the screenshot and handle press/release states correctly.
class MenuDrawerButton(ButtonBehavior, MDBoxLayout):
    """
    A custom, pill-shaped button with an icon and text, designed for menus.
    It handles its own press/release visual state to prevent a "stuck indicator".
    """

    text = StringProperty("")
    icon = StringProperty("")

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.orientation = "horizontal"
        self.size_hint_y = None
        self.height = dp(56)
        self.padding = (dp(16), 0)
        self.spacing = dp(28)

        # Define colors for normal and pressed states
        self.normal_color = (1, 1, 1, 1)  # White
        self.pressed_color = (0.9, 0.9, 0.9, 1)  # A light gray to indicate a press

        with self.canvas.before:
            self.bg_color_instruction = Color(*self.normal_color)
            self.bg_rect = RoundedRectangle(
                pos=self.pos, size=self.size, radius=[self.height / 2]
            )

        # --- Widgets inside the button ---
        self.icon_widget = MDIcon(
            icon=self.icon,
            theme_text_color="Custom",
            text_color=(0.1, 0.1, 0.1, 1),  # Dark color for the icon
            size_hint_x=None,
            width=dp(24),
            pos_hint={"center_y": 0.5},
        )
        self.label_widget = Label(
            text=self.text,
            font_size="15sp",
            color=(0.1, 0.1, 0.1, 1),  # Dark color for the text
            halign="left",
            valign="middle",
            size_hint_x=1,
        )
        self.label_widget.bind(size=self.label_widget.setter("text_size"))

        self.add_widget(self.icon_widget)
        self.add_widget(self.label_widget)

        # --- Bindings to update visuals and state ---
        self.bind(pos=self.update_graphics, size=self.update_graphics)
        self.bind(state=self.on_state_change)  # This is the key to the fix
        self.bind(
            text=self.update_text_from_property, icon=self.update_icon_from_property
        )

    def update_graphics(self, *args):
        """Updates the position and size of the background rectangle."""
        self.bg_rect.pos = self.pos
        self.bg_rect.size = self.size
        self.bg_rect.radius = [self.height / 2]

    def on_state_change(self, instance, value):
        """
        This function is called whenever the button is pressed or released.
        It changes the background color accordingly.
        """
        if value == "down":
            # Set the background to the pressed color
            self.bg_color_instruction.rgba = self.pressed_color
        else:  # value == 'normal'
            # Reset the background to the normal color
            self.bg_color_instruction.rgba = self.normal_color

    def update_text_from_property(self, instance, value):
        """Updates the label text when the `text` property changes."""
        self.label_widget.text = value

    def update_icon_from_property(self, instance, value):
        """Updates the icon graphic when the `icon` property changes."""
        self.icon_widget.icon = value


# --- CORRECTED: AppNavigationDrawer with top-aligned content ---
class AppNavigationDrawer(MDNavigationDrawer):
    def __init__(self, parent_screen, **kwargs):
        super().__init__(**kwargs)
        self.parent_screen = parent_screen

        # Set the sidebar background to a light gray color
        self.md_bg_color = (0.9, 0.9, 0.9, 1)  # Light Gray (R, G, B, A)
        self.anchor = "right"

        # This container will hold all the drawer's content.
        # The drawer itself is a vertical BoxLayout.
        drawer_content_container = MDBoxLayout(orientation="vertical")

        # Use a layout to hold our new custom buttons.
        # adaptive_height ensures this layout only takes up the vertical space it needs.
        menu_buttons_layout = MDBoxLayout(
            orientation="vertical",
            adaptive_height=True,
            spacing=dp(10),
            padding=(
                dp(10),
                dp(40),
                dp(10),
                dp(10),
            ),  # Add padding: left, top, right, bottom
        )

        # Menu Items using the new MenuDrawerButton
        profile_item = MenuDrawerButton(
            text="View Profile", icon="account-circle-outline"
        )
        profile_item.bind(on_release=self.go_to_profile)
        menu_buttons_layout.add_widget(profile_item)

        feedback_item = MenuDrawerButton(
            text="Send Feedback", icon="email-fast-outline"
        )
        feedback_item.bind(on_release=self.send_feedback)
        menu_buttons_layout.add_widget(feedback_item)

        signout_item = MenuDrawerButton(text="Sign Out", icon="logout")
        signout_item.bind(on_release=self.confirm_sign_out)
        menu_buttons_layout.add_widget(signout_item)

        # --- THE FIX: Aligning content to the top ---
        # 1. Add the menu buttons layout to the main drawer container.
        drawer_content_container.add_widget(menu_buttons_layout)

        # 2. Add a flexible spacer widget. This widget will expand to fill all
        #    the remaining vertical space, pushing the menu_buttons_layout to the top.
        drawer_content_container.add_widget(Widget())

        # 3. Add the container with the correctly aligned content to the Navigation Drawer.
        self.add_widget(drawer_content_container)

    def go_to_profile(self, *args):
        self.set_state("close")
        self.parent_screen.manager.current = "profile"

    def send_feedback(self, *args):
        self.set_state("close")
        self.parent_screen.open_feedback_popup(self.parent_screen.user_role)

    def confirm_sign_out(self, *args):
        self.set_state("close")
        self.parent_screen.confirm_sign_out()


# ########################################################################
# ####################### END: MODIFIED CODE BLOCK #######################
# ########################################################################


# New SwipeBehavior class to detect horizontal swipes
class SwipeBehavior:
    def on_touch_down(self, touch):
        # Store the starting touch position.
        self._swipe_x = touch.x
        self._swipe_y = touch.y
        # Also call the original on_touch_down event
        return super().on_touch_down(touch)

    def on_touch_up(self, touch):
        # Calculate the distance of the swipe
        dx = touch.x - self._swipe_x
        dy = touch.y - self._swipe_y

        # Define a threshold for swipe detection
        swipe_threshold = dp(100)

        # Check for a predominantly horizontal swipe.
        if abs(dx) > abs(dy) and abs(dx) > swipe_threshold:
            if dx > 0:
                # Swipe Right
                self.handle_swipe_right()
            else:
                # Swipe Left
                self.handle_swipe_left()

        # Also call the original on_touch_up event
        return super().on_touch_up(touch)

    def handle_swipe_left(self):
        """This method is intended to be overridden by classes using this behavior."""
        pass

    def handle_swipe_right(self):
        """This method is intended to be overridden by classes using this behavior."""
        pass


# Define HeaderBar class
class HeaderBar(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(size_hint=(1, None), height=120, pos_hint={"top": 1}, **kwargs)
        header_bg = Image(
            source="Images/bg_app.jpg",
            fit_mode="fill",
            size_hint=(1, 1),
            pos_hint={"x": 0, "y": 0},
        )
        self.add_widget(header_bg)


class RoundedButton(Button):
    def __init__(self, **kwargs):
        bg_color = kwargs.pop("bg_color", (1, 1, 1, 1))
        text_color = kwargs.pop("color", (0.5, 0, 0, 1))
        radius = kwargs.pop("radius", [15])

        super().__init__(**kwargs)

        self.background_normal = ""
        self.background_color = (0, 0, 0, 0)
        self.color = text_color
        self.radius = radius

        with self.canvas.before:
            self.bg = Color(*bg_color)
            self.rect = RoundedRectangle(
                pos=self.pos, size=self.size, radius=self.radius
            )

        self.bind(pos=self.update_bg, size=self.update_bg)

    def update_bg(self, *args):
        self.rect.pos = self.pos
        self.rect.size = self.size


# --- NEW: Outlined Button for Profile Screen ---
class OutlinedButton(Button):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.background_color = (0, 0, 0, 0)  # Transparent background
        self.color = (0.9, 0.9, 0.9, 1)  # White-ish text
        with self.canvas.before:
            Color(0.7, 0.7, 0.7, 1)  # Gray outline color
            self.line = Line(width=1.2)
        self.bind(pos=self.update_graphics, size=self.update_graphics)

    def update_graphics(self, *args):
        self.line.rounded_rectangle = (self.x, self.y, self.width, self.height, 24)


# --- NEW: Circular Image Widget for Profile Picture ---
class CircularImage(FloatLayout):
    source = StringProperty(None)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.img = AsyncImage(
            size_hint=(None, None),
            size=("110dp", "110dp"),
            pos_hint={"center_x": 0.5, "center_y": 0.5},
        )
        self.add_widget(self.img)
        self.bind(source=self.update_source)

        with self.canvas.before:
            StencilPush()
            self.ellipse = Ellipse(pos=self.img.pos, size=self.img.size)
            StencilUse()

        with self.canvas.after:
            StencilPop()

        self.bind(pos=self.update_stencil, size=self.update_stencil)
        self.img.bind(pos=self.update_stencil, size=self.update_stencil)

    def update_source(self, instance, value):
        self.img.source = value

    def update_stencil(self, *args):
        self.ellipse.pos = self.img.pos
        self.ellipse.size = self.img.size


# --- NEW --- Custom card for subjects with an icon
class SubjectCard(ButtonBehavior, BoxLayout):
    def __init__(self, text, **kwargs):
        super().__init__(**kwargs)
        self.orientation = "horizontal"
        self.spacing = dp(10)
        self.padding = (dp(15), dp(10))

        # Add the rounded white background
        with self.canvas.before:
            Color(1, 1, 1, 1)  # White background
            self.rect = RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(15)])

        self.bind(pos=self.update_graphics, size=self.update_graphics)

        # 1. Add the icon on the left
        icon = MDIcon(
            icon="book-open-page-variant-outline",
            font_size=dp(90),
            size_hint_x=None,
            width=dp(100),
            pos_hint={"center_y": 0.5},
            theme_text_color="Custom",
            text_color=(0.5, 0, 0, 1),
        )
        self.add_widget(icon)

        # 2. Add the label for the subject details
        details_label = Label(
            text=text,
            markup=True,
            color=(0, 0, 0, 1),
            font_size=dp(12),
            halign="left",
            valign="center",
        )
        details_label.bind(
            size=lambda *args: setattr(
                details_label, "text_size", (details_label.width, None)
            )
        )
        self.add_widget(details_label)

    def update_graphics(self, *args):
        self.rect.pos = self.pos
        self.rect.size = self.size


class CustomSpinnerOption(SpinnerOption):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.background_normal = ""
        self.background_color = (1, 1, 1, 1)
        self.color = (0.5, 0, 0, 1)
        self.font_size = 16


class ImageButton(ButtonBehavior, Image):
    pass


# QR Scanner Helper Widgets
class ScanLine(Widget):
    def __init__(self, image_widget, **kwargs):
        super().__init__(**kwargs)
        self.image_widget = image_widget
        self.y_pos = 0
        self.direction = 1
        self.is_animating = False
        with self.canvas:
            Color(0, 0.7, 1, 1)
            self.line = Line(points=[0, 0, 0, 0], width=2)
        self.image_widget.bind(size=self.update_size, pos=self.update_size)

    def update_size(self, *args):
        self.size = self.image_widget.size
        self.pos = self.image_widget.pos
        self.y_pos = self.size[1]

    def start_animation(self):
        if not self.is_animating:
            self.is_animating = True
            Clock.schedule_interval(self.update_line, 1.0 / 60)

    def stop_animation(self):
        if self.is_animating:
            self.is_animating = False
            Clock.unschedule(self.update_line)
            self.canvas.clear()

    def update_line(self, dt):
        self.y_pos -= self.direction * 3
        if self.y_pos <= 0 or self.y_pos >= self.size[1]:
            self.direction *= -1
        self.canvas.clear()
        with self.canvas:
            Color(0, 0.7, 1, 1)
            self.line = Line(
                points=[
                    self.x,
                    self.y + self.y_pos,
                    self.x + self.size[0],
                    self.y + self.y_pos,
                ],
                width=2,
            )


class CornerMarkers(Widget):
    def __init__(self, image_widget, **kwargs):
        super().__init__(**kwargs)
        self.image_widget = image_widget
        self.size_hint = (None, None)
        self.image_widget.bind(size=self.update_corners, pos=self.update_corners)
        self.update_corners()

    def update_corners(self, *args):
        self.size = self.image_widget.size
        self.pos = self.image_widget.pos
        self.canvas.clear()
        with self.canvas:
            Color(1, 1, 1, 0.8)
            marker_size = min(self.size[0], self.size[1]) * 0.15
            line_width = 4
            Line(
                points=[
                    self.x + marker_size,
                    self.y + self.size[1],
                    self.x,
                    self.y + self.size[1],
                    self.x,
                    self.y + self.size[1] - marker_size,
                ],
                width=line_width,
                cap="square",
            )
            Line(
                points=[
                    self.x + self.size[0] - marker_size,
                    self.y + self.size[1],
                    self.x + self.size[0],
                    self.y + self.size[1],
                    self.x + self.size[0],
                    self.y + self.size[1] - marker_size,
                ],
                width=line_width,
                cap="square",
            )
            Line(
                points=[
                    self.x + marker_size,
                    self.y,
                    self.x,
                    self.y,
                    self.x,
                    self.y + marker_size,
                ],
                width=line_width,
                cap="square",
            )
            Line(
                points=[
                    self.x + self.size[0] - marker_size,
                    self.y,
                    self.x + self.size[0],
                    self.y,
                    self.x + self.size[0],
                    self.y + marker_size,
                ],
                width=line_width,
                cap="square",
            )


class ModernButton(Button):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.background_color = (0, 0, 0, 0)
        self.bind(pos=self.update_graphics, size=self.update_graphics)
        with self.canvas.before:
            Color(0.8, 0.2, 0.2, 1)
            self.rect = RoundedRectangle(pos=self.pos, size=self.size, radius=[25])

    def update_graphics(self, *args):
        self.rect.pos = self.pos
        self.rect.size = self.size


# ##########################################################################
# ############ START: NEW KIVYMD NAVIGATION BAR WIDGETS ####################
# ##########################################################################
class NavigationItem(MDBoxLayout):
    """
    A custom KivyMD widget representing a single item in the navigation bar,
    consisting only of an icon.
    """

    icon_name = StringProperty()
    item_name = StringProperty()
    is_active = BooleanProperty(False)
    parent_bar = ObjectProperty(None)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.orientation = "vertical"
        self.spacing = 0
        self.padding = (0, 0, 0, 0)
        self.size_hint_y = None
        self.height = dp(50)
        self.ripple_behavior = True

        self.icon_button = MDIconButton(
            icon=self.icon_name,
            halign="center",
            valign="center",
            size_hint=(1, 1),
            theme_text_color="Custom",
            icon_size=dp(30),
        )
        self.icon_button.bind(icon=self._update_icon)
        self.icon_button.bind(on_release=self._on_button_release)

        self.add_widget(self.icon_button)
        self.bind(is_active=self._update_colors)
        self.bind(icon_name=self._update_icon)
        self._update_colors(self, self.is_active)

    def _update_icon(self, instance, value):
        self.icon_button.icon = value

    def _update_colors(self, instance, value):
        if value:  # is_active is True
            # Set the icon color to maroon when active
            self.icon_button.text_color = (0.5, 0, 0, 1)
        else:  # is_active is False
            # Set the icon color to gray when inactive
            self.icon_button.text_color = (0.5, 0.5, 0.5, 1)

    def _on_button_release(self, instance):
        if self.parent_bar:
            self.parent_bar.set_active_item(self.item_name)


# ########################################################################
# #################### START: CORRECTED NAVIGATION BAR ###################
# ########################################################################
class NavigationBar(MDBoxLayout):
    """
    The main navigation bar container. It dispatches events when items are clicked.
    """

    __events__ = ("on_home", "on_add", "on_menu")

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.orientation = "horizontal"
        self.size_hint_y = None
        self.height = dp(50)
        self.md_bg_color = (1, 1, 1, 1)  # White background
        self.elevation = 4
        self.padding = (dp(8), 0, dp(8), 0)
        self.active_item = "Home"

        self.home_item = NavigationItem(
            icon_name="home", item_name="Home", is_active=True, parent_bar=self
        )
        self.add_item = NavigationItem(
            icon_name="plus", item_name="Add", is_active=False, parent_bar=self
        )
        self.menu_item = NavigationItem(
            icon_name="menu", item_name="Menu", is_active=False, parent_bar=self
        )

        self.add_widget(self.home_item)
        self.add_widget(self.add_item)
        self.add_widget(self.menu_item)

    def update_active_item_visuals(self, item_name):
        """
        Updates which item is visually active without dispatching an event.
        This is used by screens in 'on_pre_enter' to sync the nav bar state.
        """
        self.active_item = item_name
        for item in self.children:
            if isinstance(item, NavigationItem):
                item.is_active = item.item_name == item_name

    def set_active_item(self, item_name):
        """Updates which item is currently active and dispatches an event."""
        # Special cases for re-clicking an icon (e.g., to open menu)
        if self.active_item == item_name:
            if item_name == "Menu":
                self.dispatch("on_menu")
            elif item_name == "Home":
                self.dispatch("on_home")
            return

        # Update the visuals and internal state
        self.update_active_item_visuals(item_name)

        # Dispatch the corresponding event for navigation
        if item_name == "Home":
            self.dispatch("on_home")
        elif item_name == "Add":
            self.dispatch("on_add")
        elif item_name == "Menu":
            self.dispatch("on_menu")

    def on_home(self, *args):
        pass

    def on_add(self, *args):
        pass

    def on_menu(self, *args):
        pass


# ########################################################################
# #################### END: CORRECTED NAVIGATION BAR #####################
# ########################################################################


class OutlinedSpinner(Spinner):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.background_normal = ""
        self.background_down = ""
        self.background_color = (0, 0, 0, 0)
        self.color = (0.9, 0.9, 0.9, 1)

        with self.canvas.before:
            self.border_color = Color(0.4, 0.4, 0.4, 1)
            self.border_rect = RoundedRectangle(
                pos=self.pos, size=self.size, radius=[dp(8)]
            )
            self.bg_color_instruction = Color(0.25, 0.25, 0.25, 1)
            self.bg_rect = RoundedRectangle(
                pos=(self.pos[0] + dp(1.5), self.pos[1] + dp(1.5)),
                size=(self.size[0] - dp(3), self.size[1] - dp(3)),
                radius=[dp(7)],
            )
        self.bind(
            pos=self.update_graphics, size=self.update_graphics, state=self.on_state
        )

    def update_graphics(self, *args):
        self.border_rect.pos = self.pos
        self.border_rect.size = self.size
        self.bg_rect.pos = (self.pos[0] + dp(1.5), self.pos[1] + dp(1.5))
        self.bg_rect.size = (self.size[0] - dp(3), self.size[1] - dp(3))

    def on_state(self, instance, value):
        if value == "down":
            self.bg_color_instruction.rgba = (0.3, 0.3, 0.3, 1)
            self.border_color.rgba = (0.5, 0, 0, 1)
        else:
            self.bg_color_instruction.rgba = (0.25, 0.25, 0.25, 1)
            self.border_color.rgba = (0.4, 0.4, 0.4, 1)


class QRScannerWidget(FloatLayout):
    def __init__(self, on_success_callback, on_cancel_callback, **kwargs):
        super().__init__(**kwargs)
        self.on_success = on_success_callback
        self.on_cancel = on_cancel_callback

        with self.canvas.before:
            Color(0.1, 0.1, 0.1, 1)
            self.bg = Rectangle(size=Window.size, pos=(0, 0))

        viewfinder_size = min(Window.width, Window.height) * 0.8
        self.image_container = RelativeLayout(
            size=(viewfinder_size, viewfinder_size),
            size_hint=(None, None),
            pos_hint={"center_x": 0.5, "center_y": 0.5},
        )
        self.image = Image(size_hint=(1, 1), fit_mode="cover")
        self.scan_line = ScanLine(image_widget=self.image, size_hint=(None, None))
        self.corner_markers = CornerMarkers(image_widget=self.image)
        self.image_container.add_widget(self.image)
        self.image_container.add_widget(self.scan_line)
        self.image_container.add_widget(self.corner_markers)

        self.result_label = Label(
            text="Align QR Code within the frame to scan",
            size_hint=(0.9, 0.15),
            pos_hint={"center_x": 0.5, "top": 0.9},
            font_size="18sp",
            halign="center",
            valign="middle",
            text_size=(Window.width * 0.85, None),
        )
        self.cancel_button = ModernButton(
            text="Cancel",
            size_hint=(0.6, 0.08),
            pos_hint={"center_x": 0.5, "y": 0.1},
            font_size="16sp",
        )
        self.cancel_button.bind(on_press=self.cancel_scan)

        self.add_widget(self.image_container)
        self.add_widget(self.result_label)
        self.add_widget(self.cancel_button)

        self.is_scanning = False
        self.is_camera_active = False
        self.capture = None
        Clock.schedule_once(self.start_camera, 0.5)

    def start_camera(self, dt):
        if self.is_camera_active:
            return
        self.is_camera_active = True
        self.is_scanning = True
        self.scan_line.start_animation()
        try:
            backend = cv2.CAP_DSHOW if platform == "win" else cv2.CAP_ANY
            self.capture = cv2.VideoCapture(0, backend)
            if not self.capture.isOpened():
                self.result_label.text = "Error: Could not open camera."
                self.stop_camera()
                return
            Clock.schedule_interval(self.update_camera_feed, 1.0 / 30.0)
        except Exception as e:
            self.result_label.text = f"Camera Error: {str(e)}"
            self.stop_camera()

    # ######################################################################
    # ###################### START: CORRECTED METHOD #######################
    # ######################################################################
    def update_camera_feed(self, dt):
        if not self.is_camera_active or not self.capture:
            return

        ret, frame = self.capture.read()
        if not ret:
            self.result_label.text = "Error: Failed to capture frame."
            self.stop_camera()
            return

        try:
            # --- FIX: Decode the RAW, UNFLIPPED frame first for best accuracy. ---
            if self.is_scanning:
                # 1. Convert the original frame to grayscale for pyzbar.
                gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                qr_codes = pyzbar.decode(gray_frame)

                # 2. If a code is found, process it and stop scanning.
                for qr_code in qr_codes:
                    qr_data = qr_code.data.decode("utf-8")
                    self.result_label.text = "QR Code Found!"
                    self.stop_camera()
                    self.on_success(qr_data)
                    return  # Exit the function immediately after a successful scan.

            # --- VISUALS: Prepare the frame for display AFTER attempting to decode. ---
            # 3. Flip the frame vertically. This is often needed because Kivy's
            #    texture coordinates are inverted compared to OpenCV's array coordinates.
            flipped_frame = cv2.flip(frame, 0)

            # 4. Convert the *flipped* frame to RGB for the Kivy texture.
            frame_rgb = cv2.cvtColor(flipped_frame, cv2.COLOR_BGR2RGB)

            # 5. Create and update the texture to display the (correctly oriented) image.
            texture = Texture.create(
                size=(frame.shape[1], frame.shape[0]), colorfmt="rgb"
            )
            texture.blit_buffer(frame_rgb.tobytes(), colorfmt="rgb", bufferfmt="ubyte")
            self.image.texture = texture

        except Exception as e:
            self.result_label.text = f"Error processing frame: {str(e)}"
            self.stop_camera()

    # ######################################################################
    # ####################### END: CORRECTED METHOD ########################
    # ######################################################################

    def cancel_scan(self, instance):
        self.stop_camera()
        self.on_cancel()

    def stop_camera(self, *args):
        self.is_scanning = False
        self.is_camera_active = False
        self.scan_line.stop_animation()
        Clock.unschedule(self.update_camera_feed)
        if self.capture:
            self.capture.release()
            self.capture = None


def send_feedback(sender_id, sender_role, message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = sqlite3.connect("presko.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO feedback (sender_id, sender_role, message, timestamp)
        VALUES (?, ?, ?, ?)
    """,
        (sender_id, sender_role, message, timestamp),
    )
    conn.commit()
    conn.close()


class WelcomeScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.layout = FloatLayout()

        background = Image(source="Images/bg_app.jpg", fit_mode="fill")
        self.layout.add_widget(background)

        header_bar = HeaderBar()
        self.layout.add_widget(header_bar)

        self.layout.add_widget(
            Image(
                source="Images/preskokooko-letter.png",
                size_hint=(None, None),
                size=(410, 410),
                pos_hint={"center_x": 0.5, "top": 1.05},
            )
        )

        greeting = Label(
            text="[color=ffffff]Greetings, Iskolar ng Bayan![/color]",
            markup=True,
            font_size="22sp",
            pos_hint={"center_x": 0.5, "center_y": 0.72},
        )

        student_btn = RoundedButton(
            text="Student",
            font_size=20,
            size_hint=(0.4, 0.08),
            pos_hint={"center_x": 0.5, "center_y": 0.55},
        )
        faculty_btn = RoundedButton(
            text="Faculty",
            font_size=20,
            size_hint=(0.4, 0.08),
            pos_hint={"center_x": 0.5, "center_y": 0.45},
        )
        admin_btn = RoundedButton(
            text="Admin",
            font_size=20,
            size_hint=(0.4, 0.08),
            pos_hint={"center_x": 0.5, "center_y": 0.35},
        )
        signup_btn = Button(
            text="Don't have an account? [b][color=3399FF]Sign Up[/color][/b]",
            markup=True,
            font_size=16,
            size_hint=(0.6, 0.08),
            pos_hint={"center_x": 0.5, "center_y": 0.15},
            background_color=(0, 0, 0, 0),
        )

        student_btn.bind(on_press=self.go_student)
        faculty_btn.bind(on_press=self.go_faculty)
        signup_btn.bind(on_press=lambda x: setattr(self.manager, "current", "signup"))
        admin_btn.bind(on_press=self.go_admin)

        self.layout.add_widget(greeting)
        self.layout.add_widget(student_btn)
        self.layout.add_widget(faculty_btn)
        self.layout.add_widget(signup_btn)
        self.layout.add_widget(admin_btn)
        self.add_widget(self.layout)

    def go_student(self, instance):
        self.manager.get_screen("login").set_user_type("Student")
        self.manager.current = "login"

    def go_faculty(self, instance):
        self.manager.get_screen("login").set_user_type("Professor")
        self.manager.current = "login"

    def go_admin(self, instance):
        self.manager.get_screen("login").set_user_type("Admin")
        self.manager.current = "login"


# Reusable card widget for forms
class BigRoundedCard(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.orientation = "vertical"
        self.size_hint = (None, None)
        with self.canvas.before:
            Color(1, 1, 1, 0.9)
            self.bg = RoundedRectangle(radius=[25], pos=self.pos, size=self.size)
        self.bind(pos=self.update_bg, size=self.update_bg)

    def update_bg(self, *args):
        self.bg.pos = self.pos
        self.bg.size = self.size


class SignUpScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.layout = FloatLayout()
        background = Image(source="Images/bg_app.jpg", fit_mode="fill")
        self.layout.add_widget(background)
        header_bar = HeaderBar()
        self.layout.add_widget(header_bar)
        self.layout.add_widget(
            Image(
                source="Images/preskokooko-letter.png",
                size_hint=(None, None),
                size=(380, 380),
                pos_hint={"center_x": 0.5, "top": 1.05},
            )
        )
        greeting = Label(
            text="[color=ffffff]Greetings, Iskolar ng Bayan![/color]",
            markup=True,
            font_size="22sp",
            pos_hint={"center_x": 0.5, "center_y": 0.75},
        )
        self.layout.add_widget(greeting)
        content_card = BigRoundedCard(
            size=(dp(300), dp(420)),
            pos_hint={"center_x": 0.5, "center_y": 0.37},
            padding=[dp(20), dp(20), dp(20), dp(20)],
            spacing=dp(10),
        )
        create_account_label = Label(
            text="[b]Create an Account[/b]",
            markup=True,
            font_size="24sp",
            color=(0.5, 0, 0, 1),
            size_hint_y=None,
            height=dp(40),
        )
        content_card.add_widget(create_account_label)
        self.name_input = TextInput(
            hint_text="Full Name",
            multiline=False,
            size_hint_y=None,
            height=dp(40),
            padding=[10, 10],
        )
        self.id_input = TextInput(
            hint_text="PUP ID Number",
            multiline=False,
            size_hint_y=None,
            height=dp(40),
            padding=[10, 10],
        )
        self.password_input = TextInput(
            hint_text="Password",
            password=True,
            multiline=False,
            size_hint_y=None,
            height=dp(40),
            padding=[10, 10],
        )
        self.confirm_input = TextInput(
            hint_text="Re-enter Password",
            password=True,
            multiline=False,
            size_hint_y=None,
            height=dp(40),
            padding=[10, 10],
        )
        self.role_spinner = OutlinedSpinner(
            text="Select Role",
            values=("Student", "Professor"),
            size_hint_y=None,
            height=dp(40),
            font_size=16,
            option_cls=CustomSpinnerOption,
        )
        self.inputs = [
            self.name_input,
            self.id_input,
            self.password_input,
            self.confirm_input,
        ]
        Window.bind(on_key_down=self.on_tab_key)
        signup_btn = RoundedButton(
            text="SIGN UP",
            font_size=18,
            size_hint_y=None,
            height=dp(45),
            bg_color=(0.5, 0, 0, 1),
            color=(1, 1, 1, 1),
        )
        signup_btn.bind(on_press=self.register_user)
        self.msg = Label(
            size_hint_y=None, height=dp(30), color=(0.8, 0.2, 0.2, 1), bold=True
        )
        content_card.add_widget(self.name_input)
        content_card.add_widget(self.id_input)
        content_card.add_widget(self.password_input)
        content_card.add_widget(self.confirm_input)
        content_card.add_widget(self.role_spinner)
        content_card.add_widget(signup_btn)
        content_card.add_widget(self.msg)
        self.layout.add_widget(content_card)
        back_btn = ImageButton(
            source="Images/back-button.png",
            size_hint=(None, None),
            size=(45, 45),
            pos_hint={"top": 0.98, "x": 0.01},
        )
        back_btn.bind(on_release=lambda x: setattr(self.manager, "current", "welcome"))
        self.layout.add_widget(back_btn)
        self.add_widget(self.layout)

    def register_user(self, instance):
        name = self.name_input.text.strip()
        uname = self.id_input.text.strip()
        pword = self.password_input.text.strip()
        confirm = self.confirm_input.text.strip()
        role = self.role_spinner.text.strip()

        if (
            not name
            or not uname
            or not pword
            or not confirm
            or role not in ["Student", "Professor"]
        ):
            self.msg.text = "All fields are required."
            return

        if pword != confirm:
            self.msg.text = "Passwords do not match."
            return

        try:
            with sqlite3.connect("presko.db") as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "INSERT INTO users (name, username, password, role) VALUES (?, ?, ?, ?)",
                    (name, uname, pword, role),
                )
            self.msg.text = "Registered successfully!"

            def go_to_login(dt):
                self.manager.get_screen("login").set_user_type(None)
                self.manager.current = "welcome"

            Clock.schedule_once(go_to_login, 1.5)

        except sqlite3.IntegrityError:
            self.msg.text = "User already exists."

    def on_pre_enter(self):
        self.name_input.text = ""
        self.id_input.text = ""
        self.password_input.text = ""
        self.confirm_input.text = ""
        self.role_spinner.text = "Select Role"
        self.msg.text = ""

    def on_tab_key(self, window, key, scancode, codepoint, modifiers):
        if key == 9:
            for i, field in enumerate(self.inputs):
                if field.focus:
                    next_index = (i + 1) % len(self.inputs)
                    self.inputs[next_index].focus = True
                    return True
            return False


class LoginScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.user_type = None
        self.layout = FloatLayout()
        background = Image(source="Images/bg_app.jpg", fit_mode="fill")
        self.layout.add_widget(background)
        header_bar = HeaderBar()
        self.layout.add_widget(header_bar)
        self.layout.add_widget(
            Image(
                source="Images/preskokooko-letter.png",
                size_hint=(None, None),
                size=(380, 380),
                pos_hint={"center_x": 0.5, "top": 1.05},
            )
        )
        greeting = Label(
            text="[color=ffffff]Greetings, Iskolar ng Bayan![/color]",
            markup=True,
            font_size="22sp",
            pos_hint={"center_x": 0.5, "center_y": 0.74},
        )
        self.layout.add_widget(greeting)
        self.content_card = BigRoundedCard(
            size=(dp(300), dp(310)),
            pos_hint={"center_x": 0.5, "center_y": 0.43},
            padding=[dp(20), dp(0), dp(20), dp(0)],
            spacing=dp(15),
        )
        self.title_label = Label(
            text="",
            markup=True,
            font_size="24sp",
            color=(0.5, 0, 0, 1),
            bold=True,
            size_hint_y=None,
            height=dp(40),
        )
        self.content_card.add_widget(self.title_label)
        self.username = TextInput(
            hint_text="ID Number:",
            multiline=False,
            size_hint=(1, None),
            height=dp(40),
            padding=[dp(15), dp(10), dp(15), dp(10)],
        )
        self.password = TextInput(
            hint_text="Password",
            password=True,
            multiline=False,
            size_hint=(1, None),
            height=dp(40),
            padding=[dp(15), dp(10), dp(15), dp(10)],
        )
        self.inputs = [self.username, self.password]
        Window.bind(on_key_down=self.on_tab_key)
        login_btn = RoundedButton(
            text="LOG IN",
            size_hint=(1, None),
            height=dp(45),
            bg_color=(0.5, 0, 0, 1),
            color=(1, 1, 1, 1),
            font_size=18,
        )
        login_btn.bind(on_press=self.login_user)
        forgot_btn = Button(
            text="Forgot Password?",
            font_size=14,
            size_hint_y=None,
            height=dp(30),
            background_color=(0, 0, 0, 0),
            color=(0.2, 0.2, 0.2, 1),
        )
        forgot_btn.bind(
            on_release=lambda x: setattr(self.manager, "current", "forgot_password")
        )
        self.msg = Label(
            size_hint_y=None, height=dp(30), color=(0.8, 0.2, 0.2, 1), bold=True
        )
        self.content_card.add_widget(self.username)
        self.content_card.add_widget(self.password)
        self.content_card.add_widget(login_btn)
        self.content_card.add_widget(forgot_btn)
        self.content_card.add_widget(self.msg)
        self.layout.add_widget(self.content_card)
        back_btn = ImageButton(
            source="Images/back-button.png",
            size_hint=(None, None),
            size=(45, 45),
            pos_hint={"top": 0.98, "x": 0.01},
        )
        back_btn.bind(on_release=lambda x: setattr(self.manager, "current", "welcome"))
        self.layout.add_widget(back_btn)
        self.add_widget(self.layout)

    def login_user(self, instance):
        uname = self.username.text.strip()
        pword = self.password.text.strip()
        role = self.user_type

        if not role:
            self.msg.text = "Role not selected. Please go back."
            return

        conn = sqlite3.connect("presko.db")
        cursor = conn.cursor()

        cursor.execute(
            "SELECT * FROM users WHERE username=? AND password=? AND role=?",
            (uname, pword, role),
        )
        user = cursor.fetchone()
        conn.close()

        if user:
            app = App.get_running_app()
            app.current_user_id = uname
            app.current_user_role = role  # Store role
            app.current_user_name = user[1]  # Store name

            if role == "Professor":
                app.current_professor_name = user[1]
                self.manager.current = "professor_home"
            elif role == "Student":
                app.current_student_id = uname
                app.current_student_name = user[1]
                self.manager.current = "student_home"
            elif role == "Admin":
                self.manager.current = "admin_home"

            self.msg.text = f"Login successful! ({role})"
        else:
            self.msg.text = "Invalid login credentials for selected role."

    def set_user_type(self, user_type):
        self.user_type = user_type
        if user_type == "Student":
            self.username.hint_text = "Student NO:"
            self.title_label.text = "[b]Student Login[/b]"
        elif user_type == "Professor":
            self.username.hint_text = "Professor NO:"
            self.title_label.text = "[b]Professor Login[/b]"
        elif user_type == "Admin":
            self.username.hint_text = "Admin ID:"
            self.title_label.text = "[b]Admin Login[/b]"
        else:
            self.username.hint_text = "ID Number:"
            self.title_label.text = ""

    def on_pre_enter(self):
        self.username.text = ""
        self.password.text = ""
        self.msg.text = ""

    def on_tab_key(self, window, key, scancode, codepoint, modifiers):
        if key == 9:
            for i, field in enumerate(self.inputs):
                if field.focus:
                    next_index = (i + 1) % len(self.inputs)
                    self.inputs[next_index].focus = True
                    return True
            return False


marked_dates_per_subject = {}
attendance_status_per_subject = {}
code_cache = {}
qr_expiry_cache = {}
marked_dates = {}
attendance_status = {}


# ### MODIFIED ###: This function now loads the new `late_cutoff_time` column.
def load_qr_validity_from_db():
    conn = sqlite3.connect("presko.db")
    cursor = conn.cursor()
    cursor.execute(
        "SELECT subject_code, date, manual_code, start_time, end_time, late_cutoff_time FROM qr_validity"
    )
    rows = cursor.fetchall()
    conn.close()

    for (
        subject_code,
        date,
        manual_code,
        start_time,
        end_time,
        late_cutoff_time,
    ) in rows:
        subject_key = subject_code
        full_key = f"{subject_code}_{date}"
        code_cache[full_key] = manual_code
        # Store all three timestamps in the cache for validation.
        qr_expiry_cache[full_key] = {
            "start": start_time,
            "end": end_time,
            "late_cutoff": late_cutoff_time,
        }

        if subject_key not in marked_dates_per_subject:
            marked_dates_per_subject[subject_key] = set()
        marked_dates_per_subject[subject_key].add(date)


def load_attendance_from_db():
    conn = sqlite3.connect("presko.db")
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT student_id, subject_code, date, status FROM attendance")
        rows = cursor.fetchall()
    except sqlite3.OperationalError:
        conn.close()
        return

    conn.close()

    for student_id, subject_code, date, status in rows:
        subject_key = f"{subject_code}_{student_id}"
        if subject_key not in attendance_status_per_subject:
            attendance_status_per_subject[subject_key] = {}
        attendance_status_per_subject[subject_key][date] = status


class AdminInfoCard(BoxLayout):
    def __init__(self, lines, **kwargs):
        super().__init__(
            orientation="vertical",
            spacing=5,
            padding=10,
            size_hint_y=None,
            height=120,
            **kwargs,
        )
        self.canvas.before.clear()
        with self.canvas.before:
            Color(1, 1, 1, 1)
            self.bg = RoundedRectangle(radius=[15], pos=self.pos, size=self.size)
        self.bind(pos=self.update_bg, size=self.update_bg)

        for line in lines:
            self.add_widget(
                Label(
                    text=line,
                    font_size=14,
                    color=(0, 0, 0, 1),
                    halign="left",
                    valign="middle",
                    text_size=(self.width - dp(20), None),
                    markup=True,
                )
            )

    def update_bg(self, *args):
        self.bg.pos = self.pos
        self.bg.size = self.size
        for child in self.children:
            child.text_size = (self.width - dp(20), None)


class StudentHomeScreen(Screen, SwipeBehavior):
    subjects = ListProperty([])
    user_role = StringProperty("Student")

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.layout = BoxLayout(orientation="vertical")
        self.subjects = []

        main_content_area = FloatLayout()
        bg = Image(source="Images/bg_app.jpg", fit_mode="fill")
        main_content_area.add_widget(bg)
        header_bar = HeaderBar()
        main_content_area.add_widget(header_bar)
        main_content_area.add_widget(
            Image(
                source="Images/preskokooko-letter.png",
                size_hint=(None, None),
                size=(200, 200),
                pos_hint={"center_x": 0.5, "top": 1},
            )
        )
        student_label = Label(
            text="[b]STUDENT[/b]",
            markup=True,
            font_size="18sp",
            size_hint=(0.3, 0.05),
            pos_hint={"center_x": 0.5, "top": 0.80},
            color=(1, 1, 1, 1),
            halign="center",
            valign="middle",
        )
        with student_label.canvas.before:
            Color(0.5, 0, 0, 1)
            student_bg = Rectangle(size=student_label.size, pos=student_label.pos)
        student_label.bind(
            pos=lambda i, v: setattr(student_bg, "pos", v),
            size=lambda i, v: setattr(student_bg, "size", v),
        )
        main_content_area.add_widget(student_label)
        self.datetime_label = Label(
            text="",
            size_hint=(None, None),
            size=(400, 40),
            pos_hint={"center_x": 0.5, "top": 0.97},
            font_size=18,
            color=(1, 1, 1, 1),
        )
        Clock.schedule_interval(self.update_datetime, 1)
        main_content_area.add_widget(self.datetime_label)

        scroll_container = FloatLayout(
            size_hint=(0.9, 0.6),
            pos_hint={"center_x": 0.5, "top": 0.72},
        )
        with scroll_container.canvas.before:
            Color(1, 1, 1, 0.3)
            self.scroll_bg = RoundedRectangle(
                pos=scroll_container.pos,
                size=scroll_container.size,
                radius=[dp(20)],
            )

        def update_scroll_bg(instance, value):
            self.scroll_bg.pos = instance.pos
            self.scroll_bg.size = instance.size

        scroll_container.bind(pos=update_scroll_bg, size=update_scroll_bg)

        scrollview = ScrollView(
            size_hint=(1, 1),
            pos_hint={"center_x": 0.5, "center_y": 0.5},
            bar_width=0,
        )

        self.subject_box = BoxLayout(
            orientation="vertical",
            size_hint_y=None,
            padding=(dp(10), dp(10)),  # Default padding for when list is populated
            spacing=dp(10),
        )
        self.subject_box.bind(minimum_height=self.subject_box.setter("height"))
        scrollview.add_widget(self.subject_box)

        scroll_container.add_widget(scrollview)
        main_content_area.add_widget(scroll_container)

        self.nav_bar = NavigationBar()
        self.nav_bar.bind(on_home=self.on_pre_enter)
        self.nav_bar.bind(on_add=self.go_to_add_subject)
        self.nav_bar.bind(on_menu=self.toggle_nav_drawer)

        self.nav_drawer = AppNavigationDrawer(parent_screen=self)
        main_content_area.add_widget(self.nav_drawer)

        self.layout.add_widget(main_content_area)
        self.layout.add_widget(self.nav_bar)
        self.add_widget(self.layout)

    def toggle_nav_drawer(self, *args):
        self.nav_drawer.set_state("open")

    def handle_swipe_left(self):
        self.go_to_add_subject()

    def update_datetime(self, dt):
        now = datetime.now()
        self.datetime_label.text = now.strftime("%A, %B %d, %Y | %I:%M:%S %p")

    def load_subject_cards(self):
        self.subject_box.clear_widgets()
        self.subjects = []
        student_id = App.get_running_app().current_student_id
        if not student_id:
            # If there's no student ID, show the centered message.
            self.subject_box.size_hint_y = 1
            self.subject_box.padding = 0
            self.subject_box.add_widget(Widget())
            self.subject_box.add_widget(
                Label(
                    text="No subjects added yet.\nTap the '+' button to add one.",
                    color=(0.1, 0.1, 0.1, 1),
                    halign="center",
                )
            )
            self.subject_box.add_widget(Widget())
            return

        conn = sqlite3.connect("presko.db")
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT s.subject_code, s.subject_name, s.professor_name, s.schedule, s.section
            FROM subjects s
            JOIN student_subjects ss ON s.subject_code = ss.subject_code
            WHERE ss.student_id = ?
        """,
            (student_id,),
        )
        self.subjects = cursor.fetchall()
        conn.close()

        if not self.subjects:
            # FIX: If the list is empty, fill the container and center the label.
            # 1. Make the layout fill the vertical space.
            self.subject_box.size_hint_y = 1
            # 2. Remove padding to allow for true centering.
            self.subject_box.padding = 0

            # 3. Add a top spacer widget to push the label down.
            self.subject_box.add_widget(Widget())
            # 4. Add the actual label.
            self.subject_box.add_widget(
                Label(
                    text="No subjects added yet.\nTap the '+' button to add one.",
                    color=(0.1, 0.1, 0.1, 1),
                    halign="center",
                )
            )
            # 5. Add a bottom spacer to push the label up.
            self.subject_box.add_widget(Widget())
        else:
            # FIX: If the list has items, ensure it's a normal scrollable list.
            # 1. Allow the layout to have a dynamic height for scrolling.
            self.subject_box.size_hint_y = None
            # 2. Set padding for the subject cards.
            self.subject_box.padding = (dp(10), dp(10))

            for data in self.subjects:
                text = (
                    f"[b]SUBJECT CODE:[/b] {data[0]}\n"
                    f"[b]SUBJECT NAME:[/b] {data[1]}\n"
                    f"[b]PROFESSOR:[/b] {data[2]}\n"
                    f"[b]SCHEDULE:[/b] {data[3]}\n"
                    f"[b]SECTION:[/b] {data[4]}"
                )
                card = SubjectCard(
                    text=text,
                    size_hint_y=None,
                    height=dp(120),
                )
                subject_key = f"{data[0]}_{student_id}"
                card.bind(on_release=lambda x, sk=subject_key: self.open_calendar(sk))
                self.subject_box.add_widget(card)

    def open_calendar(self, subject_key):
        calendar_screen = self.manager.get_screen("student_calendar")
        calendar_screen.current_subject_key = subject_key
        self.manager.current = "student_calendar"

    def go_to_add_subject(self, *args):
        self.manager.transition.direction = "left"
        self.manager.current = "add_subject_student"

    def confirm_sign_out(self, *args):
        content = BoxLayout(orientation="vertical", spacing=10, padding=20)
        content.add_widget(
            Label(text="Are you sure you want to sign out?", color=(1, 1, 1, 1))
        )
        btn_layout = BoxLayout(spacing=10, size_hint_y=None, height=dp(40))
        yes_btn = RoundedButton(
            text="Yes", bg_color=(0.8, 0.2, 0.2, 1), color=(1, 1, 1, 1)
        )
        no_btn = RoundedButton(text="No")
        btn_layout.add_widget(yes_btn)
        btn_layout.add_widget(no_btn)
        content.add_widget(btn_layout)
        self.confirm_popup = Popup(
            title="Confirm Sign Out",
            title_color=(1, 1, 1, 1),
            content=content,
            size_hint=(0.8, 0.3),
            background_color=(0.2, 0.2, 0.2, 0.95),
            separator_color=(0.5, 0, 0, 1),
        )
        yes_btn.bind(on_release=self.sign_out)
        no_btn.bind(on_release=self.confirm_popup.dismiss)
        self.confirm_popup.open()

    def sign_out(self, *args):
        self.confirm_popup.dismiss()
        self.manager.get_screen("login").set_user_type(None)
        self.manager.current = "welcome"

    def on_pre_enter(self, *args):
        self.nav_bar.update_active_item_visuals("Home")
        self.load_subject_cards()

    def open_feedback_popup(self, role):
        layout = BoxLayout(orientation="vertical", spacing=10, padding=20)
        input_box = TextInput(
            hint_text="Write your feedback here...", multiline=True, size_hint=(1, 0.7)
        )
        layout.add_widget(input_box)
        status_label = Label(text="", size_hint=(1, 0.1), color=(1, 1, 1, 1))
        layout.add_widget(status_label)
        submit_btn = RoundedButton(
            text="Submit Feedback",
            size_hint=(1, 0.2),
            bg_color=(0.1, 0.5, 0.1, 1),
            color=(1, 1, 1, 1),
        )
        layout.add_widget(submit_btn)
        popup = Popup(
            title="Send Feedback",
            title_color=(1, 1, 1, 1),
            content=layout,
            size_hint=(0.9, 0.6),
            background_color=(0.2, 0.2, 0.2, 0.95),
            separator_color=(0.5, 0, 0, 1),
        )

        def submit_feedback(instance):
            message = input_box.text.strip()
            if not message:
                status_label.text = "Feedback cannot be empty."
                return
            app = App.get_running_app()
            sender_id = app.current_user_id
            send_feedback(sender_id, role, message)
            status_label.text = "Feedback sent successfully!"
            Clock.schedule_once(lambda dt: popup.dismiss(), 1.5)

        submit_btn.bind(on_release=submit_feedback)
        popup.open()

    def on_leave(self, *args):
        """Closes the navigation drawer when the screen is left."""
        self.nav_drawer.set_state("close")


# ### MODIFIED ###: This entire screen is now simplified for code entry.
class AddSubjectStudentScreen(Screen, SwipeBehavior):
    user_role = StringProperty("Student")

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.layout = BoxLayout(orientation="vertical")

        main_content_area = FloatLayout()
        main_content_area.add_widget(Image(source="Images/bg_app.jpg", fit_mode="fill"))
        header_bar = HeaderBar()
        main_content_area.add_widget(header_bar)

        main_content_area.add_widget(
            Image(
                source="Images/preskokooko-letter.png",
                size_hint=(None, None),
                size=(350, 350),
                pos_hint={"center_x": 0.5, "top": 1},
            )
        )

        self.datetime_label = Label(
            text="",
            size_hint=(None, None),
            size=(400, 40),
            pos_hint={"center_x": 0.5, "top": 0.97},
            font_size=18,
            color=(1, 1, 1, 1),  # MODIFIED: Changed from black to white
        )
        main_content_area.add_widget(self.datetime_label)
        Clock.schedule_interval(self.update_datetime, 1)

        # Simplified form container
        form_container = BigRoundedCard(
            size=(dp(300), dp(200)),  # Adjusted size
            pos_hint={"center_x": 0.5, "center_y": 0.48},
            padding=dp(20),
            spacing=dp(15),
        )

        form_container.add_widget(
            Label(
                text="Add a New Subject",
                font_size="22sp",
                color=(0.5, 0, 0, 1),
                bold=True,
                size_hint_y=None,
                height=dp(30),
            )
        )

        # Single input for the subject code
        self.subject_code_input = TextInput(
            hint_text="Enter Subject Code from Professor",
            size_hint_y=None,
            height=dp(45),
            font_size=dp(14),
            multiline=False,
        )
        form_container.add_widget(self.subject_code_input)

        # Submit button
        submit_btn = RoundedButton(
            text="Add Subject",
            size_hint_y=None,
            height=dp(45),
            bg_color=(0.5, 0, 0, 1),
            color=(1, 1, 1, 1),
        )
        submit_btn.bind(on_release=self.submit)
        form_container.add_widget(submit_btn)
        main_content_area.add_widget(form_container)

        self.layout.add_widget(main_content_area)

        self.nav_bar = NavigationBar()
        self.nav_bar.bind(on_home=self.go_to_home)
        self.nav_bar.bind(on_add=lambda x: None)
        self.nav_bar.bind(on_menu=self.toggle_nav_drawer)

        self.nav_drawer = AppNavigationDrawer(parent_screen=self)
        main_content_area.add_widget(self.nav_drawer)

        self.layout.add_widget(self.nav_bar)
        self.add_widget(self.layout)

    def toggle_nav_drawer(self, *args):
        self.nav_drawer.set_state("open")

    def handle_swipe_right(self):
        self.go_to_home()

    def go_to_home(self, *args):
        self.manager.transition.direction = "right"
        self.manager.current = "student_home"

    def on_pre_enter(self, *args):
        self.nav_bar.update_active_item_visuals("Add")
        self.subject_code_input.text = ""

    def update_datetime(self, dt):
        now = datetime.now()
        self.datetime_label.text = now.strftime("%A, %B %d, %Y | %I:%M:%S %p")

    def open_feedback_popup(self, role):
        layout = BoxLayout(orientation="vertical", spacing=10, padding=20)
        input_box = TextInput(
            hint_text="Write your feedback here...", multiline=True, size_hint=(1, 0.7)
        )
        layout.add_widget(input_box)
        status_label = Label(
            text="", size_hint=(1, 0.1), color=(1, 1, 1, 1)
        )  # MODIFIED
        layout.add_widget(status_label)
        submit_btn = RoundedButton(
            text="Submit Feedback",
            size_hint=(1, 0.2),
            bg_color=(0.1, 0.5, 0.1, 1),
            color=(1, 1, 1, 1),
        )
        layout.add_widget(submit_btn)
        popup = Popup(
            title="Send Feedback",
            title_color=(1, 1, 1, 1),  # MODIFIED
            content=layout,
            size_hint=(0.9, 0.6),
            background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
            separator_color=(0.5, 0, 0, 1),
        )

        def submit_feedback(instance):
            message = input_box.text.strip()
            if not message:
                status_label.text = "Feedback cannot be empty."
                return
            app = App.get_running_app()
            sender_id = app.current_user_id
            send_feedback(sender_id, role, message)
            status_label.text = "Feedback sent successfully!"
            Clock.schedule_once(lambda dt: popup.dismiss(), 1.5)

        submit_btn.bind(on_release=submit_feedback)
        popup.open()

    def confirm_sign_out(self, *args):
        content = BoxLayout(orientation="vertical", spacing=10, padding=20)
        content.add_widget(
            Label(
                text="Are you sure you want to sign out?", color=(1, 1, 1, 1)
            )  # MODIFIED
        )
        btn_layout = BoxLayout(spacing=10, size_hint_y=None, height=dp(40))
        yes_btn = RoundedButton(
            text="Yes", bg_color=(0.8, 0.2, 0.2, 1), color=(1, 1, 1, 1)
        )
        no_btn = RoundedButton(text="No")
        btn_layout.add_widget(yes_btn)
        btn_layout.add_widget(no_btn)
        content.add_widget(btn_layout)
        self.confirm_popup = Popup(
            title="Confirm Sign Out",
            title_color=(1, 1, 1, 1),  # MODIFIED
            content=content,
            size_hint=(0.8, 0.3),
            background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
            separator_color=(0.5, 0, 0, 1),
        )
        yes_btn.bind(on_release=self.sign_out)
        no_btn.bind(on_release=self.confirm_popup.dismiss)
        self.confirm_popup.open()

    def sign_out(self, *args):
        self.confirm_popup.dismiss()
        self.manager.get_screen("login").set_user_type(None)
        self.manager.current = "welcome"

    def submit(self, instance):
        student_id = App.get_running_app().current_student_id
        subject_code = self.subject_code_input.text.strip().upper()

        if not subject_code:
            popup = Popup(
                title="Input Error",
                content=Label(
                    text="Please enter a subject code.", color=(1, 1, 1, 1)
                ),  # MODIFIED
                size_hint=(0.8, 0.3),
                background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
                title_color=(1, 1, 1, 1),  # MODIFIED
            )
            popup.open()
            return

        conn = sqlite3.connect("presko.db")
        cursor = conn.cursor()

        # Check if the subject code exists in the main subjects table
        cursor.execute("SELECT * FROM subjects WHERE subject_code=?", (subject_code,))
        subject_exists = cursor.fetchone()

        if not subject_exists:
            conn.close()
            popup = Popup(
                title="Subject Not Found",
                content=Label(
                    text="Invalid subject code.\nPlease check the code and try again.",
                    color=(1, 1, 1, 1),  # MODIFIED
                    halign="center",
                ),
                size_hint=(0.8, 0.3),
                background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
                title_color=(1, 1, 1, 1),  # MODIFIED
            )
            popup.open()
            return

        # Check if the student is already enrolled in this subject
        cursor.execute(
            "SELECT * FROM student_subjects WHERE student_id = ? AND subject_code = ?",
            (student_id, subject_code),
        )
        already_enrolled = cursor.fetchone()

        if already_enrolled:
            conn.close()
            popup = Popup(
                title="Already Enrolled",
                content=Label(
                    text="You are already enrolled in this subject.",
                    color=(1, 1, 1, 1),  # MODIFIED
                ),
                size_hint=(0.8, 0.3),
                background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
                title_color=(1, 1, 1, 1),  # MODIFIED
            )
            popup.open()
            return

        # If code is valid and student is not enrolled, add the subject
        cursor.execute(
            "INSERT INTO student_subjects (student_id, subject_code) VALUES (?, ?)",
            (student_id, subject_code),
        )
        conn.commit()
        conn.close()

        # Success popup and navigate home
        popup = Popup(
            title="Success",
            content=Label(
                text=f"Successfully added subject:\n{subject_code}",
                color=(1, 1, 1, 1),  # MODIFIED
                halign="center",
            ),
            size_hint=(0.8, 0.3),
            background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
            title_color=(1, 1, 1, 1),  # MODIFIED
        )
        popup.open()

        self.manager.current = "student_home"

    def on_leave(self, *args):
        """Closes the navigation drawer when the screen is left."""
        self.nav_drawer.set_state("close")


# ==============================================================================
# ========= START: STUDENT CALENDAR SCREEN WITH CORRECTED SCAN LOGIC ===========
# ==============================================================================
class StudentCalendarScreen(Screen, SwipeBehavior):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.layout = BoxLayout(orientation="vertical")
        self.scanner_popup = None
        self.scanner_widget = None
        self.current_subject_key = None
        self.current_year = datetime.now().year
        self.current_month = datetime.now().month
        self.selected_date = None

        main_content_area = FloatLayout()
        bg = Image(source="Images/bg_app.jpg", fit_mode="fill")
        main_content_area.add_widget(bg)
        header_bar = HeaderBar()
        main_content_area.add_widget(header_bar)

        main_content_area.add_widget(
            Image(
                source="Images/preskokooko-letter.png",
                size_hint=(None, None),
                size=(200, 200),
                pos_hint={"center_x": 0.5, "top": 0.95},
            )
        )

        self.datetime_label = Label(
            text="",
            size_hint=(None, None),
            size=(400, 40),
            pos_hint={"center_x": 0.5, "top": 0.97},
            font_size=18,
            color=(1, 1, 1, 1),
        )
        Clock.schedule_interval(self.update_datetime, 1)
        main_content_area.add_widget(self.datetime_label)

        rounded_container = BigRoundedCard(
            size=(dp(320), dp(420)),
            pos_hint={"center_x": 0.5, "center_y": 0.42},
            padding=[dp(20), dp(10), dp(20), dp(20)],
            spacing=dp(5),
        )

        nav_buttons = BoxLayout(size_hint_y=None, height=dp(40), spacing=5)
        self.month_label = Label(
            text=f"{calendar.month_name[self.current_month]} {self.current_year}",
            font_size=dp(16),
            color=(0.5, 0, 0, 1),
            bold=True,
        )
        prev_btn = Button(
            text="<",
            size_hint_x=0.2,
            background_color=(0.8, 0.8, 0.8, 1),
            color=(0, 0, 0, 1),
        )
        next_btn = Button(
            text=">",
            size_hint_x=0.2,
            background_color=(0.8, 0.8, 0.8, 1),
            color=(0, 0, 0, 1),
        )
        prev_btn.bind(on_release=self.go_to_prev_month)
        next_btn.bind(on_release=self.go_to_next_month)
        nav_buttons.add_widget(prev_btn)
        nav_buttons.add_widget(self.month_label)
        nav_buttons.add_widget(next_btn)
        rounded_container.add_widget(nav_buttons)

        self.calendar_grid = GridLayout(cols=7, spacing=dp(2))
        rounded_container.add_widget(self.calendar_grid)

        self.show_qr_btn = RoundedButton(
            text="Scan QR Code",
            size_hint_y=None,
            height=dp(40),
            bg_color=(0.2, 0.5, 0.8, 1),
            color=(1, 1, 1, 1),
            disabled=True,
        )
        self.show_qr_btn.bind(on_release=self.open_qr_scanner_popup)

        self.enter_code_btn = RoundedButton(
            text="Enter Manual Code",
            size_hint_y=None,
            height=dp(40),
            bg_color=(0.2, 0.2, 0.2, 1),
            color=(1, 1, 1, 1),
            disabled=True,
        )
        self.enter_code_btn.bind(on_release=self.open_code_entry_popup)

        rounded_container.add_widget(Widget(size_hint_y=0.1))
        rounded_container.add_widget(self.show_qr_btn)
        rounded_container.add_widget(self.enter_code_btn)

        main_content_area.add_widget(rounded_container)

        self.layout.add_widget(main_content_area)

        self.nav_bar = NavigationBar()
        self.nav_bar.bind(
            on_home=lambda x: setattr(self.manager, "current", "student_home")
        )
        self.nav_bar.bind(
            on_add=lambda x: setattr(self.manager, "current", "add_subject_student")
        )

        self.layout.add_widget(self.nav_bar)
        self.add_widget(self.layout)

    def handle_swipe_right(self):
        self.manager.transition.direction = "right"
        self.manager.current = "student_home"

    def on_pre_enter(self, *args):
        self.nav_bar.update_active_item_visuals("Home")

        marked_dates_per_subject.clear()
        attendance_status_per_subject.clear()
        code_cache.clear()
        qr_expiry_cache.clear()
        load_qr_validity_from_db()
        load_attendance_from_db()
        today = datetime.now()
        self.current_year = today.year
        self.current_month = today.month
        self.month_label.text = (
            f"{calendar.month_name[self.current_month]} {self.current_year}"
        )
        if not self.current_subject_key:
            print("No subject selected!")
            return
        self.generate_calendar(self.current_year, self.current_month)
        self.show_qr_btn.disabled = True
        self.enter_code_btn.disabled = True

    def generate_calendar(self, year, month):
        if not self.current_subject_key:
            return

        green = (0.2, 0.8, 0.2, 1)
        red = (1, 0.2, 0.2, 1)
        white = (1, 1, 1, 0.8)
        orange = (1, 0.65, 0, 1)
        yellow = (1, 1, 0.2, 1)

        self.calendar_grid.clear_widgets()
        calendar.setfirstweekday(calendar.SUNDAY)

        for day in ["S", "M", "T", "W", "T", "F", "S"]:
            self.calendar_grid.add_widget(
                Label(text=day, color=(0.5, 0, 0, 1), bold=True, font_size=dp(12))
            )

        month_days = calendar.monthcalendar(year, month)
        subject_code = self.current_subject_key.split("_")[0]

        marked_dates = marked_dates_per_subject.get(subject_code, set())
        attendance = attendance_status_per_subject.get(self.current_subject_key, {})

        for week in month_days:
            for day in week:
                if day == 0:
                    self.calendar_grid.add_widget(Widget())
                else:
                    date_str = f"{year}-{month:02d}-{day:02d}"
                    status = attendance.get(date_str)
                    color = white

                    if status == "present":
                        color = green
                    elif status == "late":
                        color = orange
                    elif status == "absent":
                        color = red
                    elif date_str in marked_dates:
                        color = yellow

                    btn = Button(
                        text=str(day),
                        background_normal="",
                        background_color=color,
                        color=(0, 0, 0, 1),
                        font_size=dp(12),
                    )
                    btn.bind(on_release=partial(self.on_date_selected, date_str))
                    self.calendar_grid.add_widget(btn)

    def update_datetime(self, dt):
        now = datetime.now()
        self.datetime_label.text = now.strftime("%A, %B %d, %Y | %I:%M:%S %p")

    def go_to_prev_month(self, instance):
        self.current_month -= 1
        if self.current_month == 0:
            self.current_month = 12
            self.current_year -= 1
        self.month_label.text = (
            f"{calendar.month_name[self.current_month]} {self.current_year}"
        )
        self.generate_calendar(self.current_year, self.current_month)

    def go_to_next_month(self, instance):
        self.current_month += 1
        if self.current_month == 13:
            self.current_month = 1
            self.current_year += 1
        self.month_label.text = (
            f"{calendar.month_name[self.current_month]} {self.current_year}"
        )
        self.generate_calendar(self.current_year, self.current_month)

    def on_date_selected(self, date_str, *args):
        self.selected_date = date_str
        subject_code = self.current_subject_key.split("_")[0]
        student_id = App.get_running_app().current_student_id
        full_key = f"{subject_code}_{student_id}"
        marked_dates = marked_dates_per_subject.get(subject_code, set())
        attendance = attendance_status_per_subject.get(full_key, {})
        if attendance.get(date_str) in ["present", "late", "absent"]:
            self.show_qr_btn.disabled = True
            self.enter_code_btn.disabled = True
        elif date_str in marked_dates:
            self.show_qr_btn.disabled = False
            self.enter_code_btn.disabled = False
        else:
            self.show_qr_btn.disabled = True
            self.enter_code_btn.disabled = True

    def open_qr_scanner_popup(self, instance):
        if not self.selected_date:
            return
        self.scanner_widget = QRScannerWidget(
            on_success_callback=self.handle_qr_scan_success,
            on_cancel_callback=self.handle_qr_scan_cancel,
        )
        self.scanner_popup = Popup(
            title="Scan QR Code",
            content=self.scanner_widget,
            size_hint=(1, 1),
            auto_dismiss=False,
            title_size=0,
            separator_height=0,
        )
        self.scanner_popup.bind(on_dismiss=self.scanner_widget.stop_camera)
        self.scanner_popup.open()

    # ######################################################################
    # #################### START: CORRECTED FUNCTION #######################
    # ######################################################################
    def handle_qr_scan_success(self, qr_data):
        """
        Handles the logic after a QR code has been successfully scanned.
        Validates the code and the time, then records attendance.
        """
        if self.scanner_popup:
            self.scanner_popup.dismiss()

        subject_code = self.current_subject_key.split("_")[0]
        key = f"{subject_code}_{self.selected_date}"
        expected_code = code_cache.get(key)

        # Compare the scanned data with the expected code from the cache
        if qr_data.strip() == expected_code:
            # Code is correct, now validate the time
            status, message = self._validate_time_and_get_status()

            # If status is None, it means time validation failed (e.g., too early)
            # or the 'absent' case was already handled internally.
            if status is None:
                # Check if a specific error message was returned to be shown
                if message != "This will be handled by the absent call.":
                    popup = Popup(
                        title="Time Error",
                        content=Label(
                            text=message, color=(1, 1, 1, 1), halign="center"
                        ),
                        size_hint=(0.8, 0.3),
                        background_color=(0.2, 0.2, 0.2, 0.95),
                        separator_color=(0.5, 0, 0, 1),
                        title_color=(1, 1, 1, 1),
                    )
                    popup.open()
                return  # Stop further processing

            # If we have a valid status ('present' or 'late'), record it.
            self.record_attendance(status, message)
        else:
            # The scanned code did not match the expected code
            self.record_attendance(None, "❌ Incorrect or expired QR code.")

    # ######################################################################
    # ##################### END: CORRECTED FUNCTION ########################
    # ######################################################################

    def _validate_time_and_get_status(self):
        subject_code = self.current_subject_key.split("_")[0]
        key = f"{subject_code}_{self.selected_date}"
        validity = qr_expiry_cache.get(key)

        if not validity:
            return None, "Error: QR code validity information not found for this day."

        try:
            now = datetime.now()
            start = datetime.strptime(validity["start"], "%Y-%m-%d %H:%M:%S")
            end = datetime.strptime(validity["end"], "%Y-%m-%d %H:%M:%S")
            late_cutoff = datetime.strptime(
                validity["late_cutoff"], "%Y-%m-%d %H:%M:%S"
            )
        except (ValueError, TypeError, KeyError):
            return (
                None,
                "Error: Invalid time format in DB. Please ask professor to regenerate QR.",
            )

        if now < start:
            return None, "⏰ QR is not yet valid. Try again later."
        elif start <= now <= late_cutoff:
            return "present", "✅ You are marked present."
        elif late_cutoff < now <= end:
            return "late", "🟠 You are marked late."
        else:
            self.record_attendance(
                "absent", "❌ Attendance window missed. Marked absent."
            )
            return (
                None,
                "This will be handled by the absent call.",
            )

    def open_code_entry_popup(self, instance):
        # This function remains the same as before
        if not self.selected_date:
            return

        content = BoxLayout(orientation="vertical", padding=20, spacing=10)
        code_input = TextInput(
            hint_text="Enter Code",
            multiline=False,
            font_size=dp(18),
            size_hint_y=None,
            height=dp(50),
            halign="center",
        )
        status_label = Label(
            text="",
            font_size=dp(16),
            size_hint_y=None,
            height=dp(30),
            color=(1, 1, 1, 1),
        )
        submit_btn = RoundedButton(text="Submit", size_hint_y=None, height=dp(50))
        content.add_widget(code_input)
        content.add_widget(status_label)
        content.add_widget(submit_btn)

        popup = Popup(
            title="Enter Manual Code",
            content=content,
            size_hint=(0.9, 0.4),
            title_color=(1, 1, 1, 1),
            separator_color=(0.5, 0, 0, 1),
            background_color=(0.2, 0.2, 0.2, 0.95),
        )

        def check_code(btn_instance):
            subject_code = self.current_subject_key.split("_")[0]
            key = f"{subject_code}_{self.selected_date}"
            expected_code = code_cache.get(key)

            if code_input.text.strip().upper() == expected_code:
                status, message = self._validate_time_and_get_status()

                if status is None:
                    if message != "This will be handled by the absent call.":
                        status_label.text = message
                    else:
                        popup.dismiss()
                    return

                self.record_attendance(status, message)
                popup.dismiss()
            else:
                status_label.text = "❌ Incorrect or expired code."

        submit_btn.bind(on_release=check_code)
        popup.open()

    def handle_qr_scan_cancel(self):
        if self.scanner_popup:
            self.scanner_popup.dismiss()

    def record_attendance(self, status, message):
        popup = Popup(
            title="Attendance Status",
            content=Label(text=message, color=(1, 1, 1, 1), halign="center"),
            size_hint=(0.8, 0.3),
            background_color=(0.2, 0.2, 0.2, 0.95),
            separator_color=(0.5, 0, 0, 1),
            title_color=(1, 1, 1, 1),
        )
        popup.open()

        if status:
            subject_key = self.current_subject_key
            if subject_key not in attendance_status_per_subject:
                attendance_status_per_subject[subject_key] = {}
            attendance_status_per_subject[subject_key][self.selected_date] = status

            subject_code, student_id = subject_key.split("_")
            with sqlite3.connect("presko.db") as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT id FROM attendance WHERE student_id=? AND subject_code=? AND date=?",
                    (student_id, subject_code, self.selected_date),
                )
                if cursor.fetchone():
                    cursor.execute(
                        "UPDATE attendance SET status=? WHERE student_id=? AND subject_code=? AND date=?",
                        (status, student_id, subject_code, self.selected_date),
                    )
                else:
                    cursor.execute(
                        "INSERT INTO attendance (student_id, subject_code, date, status) VALUES (?, ?, ?, ?)",
                        (student_id, subject_code, self.selected_date, status),
                    )

            self.generate_calendar(self.current_year, self.current_month)
            if self.selected_date:
                self.on_date_selected(self.selected_date)

        Clock.schedule_once(lambda dt: popup.dismiss(), 2)


# ==============================================================================
# ================= END: STUDENT CALENDAR SCREEN CORRECTION ====================
# ==============================================================================


class ProfessorHomeScreen(Screen, SwipeBehavior):
    subjects = ListProperty([])
    user_role = StringProperty("Professor")

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.layout = BoxLayout(orientation="vertical")
        self.subjects = []
        main_content_area = FloatLayout()
        bg = Image(source="Images/bg_app.jpg", fit_mode="fill")
        main_content_area.add_widget(bg)
        header_bar = HeaderBar()
        main_content_area.add_widget(header_bar)
        main_content_area.add_widget(
            Image(
                source="Images/preskokooko-letter.png",
                size_hint=(None, None),
                size=(200, 200),
                pos_hint={"center_x": 0.5, "top": 1},
            )
        )
        professor_label = Label(
            text="[b]PROFESSOR[/b]",
            markup=True,
            font_size="18sp",
            size_hint=(0.4, 0.05),
            pos_hint={"center_x": 0.5, "top": 0.80},
            color=(1, 1, 1, 1),
            halign="center",
            valign="middle",
        )
        with professor_label.canvas.before:
            Color(0.5, 0, 0, 1)
            self.prof_label_bg = Rectangle(
                size=professor_label.size, pos=professor_label.pos
            )
        professor_label.bind(
            pos=lambda i, v: setattr(self.prof_label_bg, "pos", v),
            size=lambda i, v: setattr(self.prof_label_bg, "size", v),
        )
        main_content_area.add_widget(professor_label)
        self.datetime_label = Label(
            text="",
            size_hint=(None, None),
            size=(400, 40),
            pos_hint={"center_x": 0.5, "top": 0.97},
            font_size=18,
            color=(1, 1, 1, 1),  # MODIFIED: Changed from black to white
        )
        Clock.schedule_interval(self.update_datetime, 1)
        main_content_area.add_widget(self.datetime_label)

        scroll_container = FloatLayout(
            size_hint=(0.9, 0.6),
            pos_hint={"center_x": 0.5, "top": 0.72},
        )
        with scroll_container.canvas.before:
            Color(1, 1, 1, 0.3)
            self.scroll_bg = RoundedRectangle(
                pos=scroll_container.pos,
                size=scroll_container.size,
                radius=[dp(20)],
            )

        def update_scroll_bg(instance, value):
            self.scroll_bg.pos = instance.pos
            self.scroll_bg.size = instance.size

        scroll_container.bind(pos=update_scroll_bg, size=update_scroll_bg)

        scrollview = ScrollView(
            size_hint=(1, 1),
            pos_hint={"center_x": 0.5, "center_y": 0.5},
            bar_width=0,
        )

        self.subject_box = BoxLayout(
            orientation="vertical",
            size_hint_y=None,
            padding=(dp(10), dp(10)),
            spacing=dp(10),
        )
        self.subject_box.bind(minimum_height=self.subject_box.setter("height"))
        scrollview.add_widget(self.subject_box)

        scroll_container.add_widget(scrollview)
        main_content_area.add_widget(scroll_container)

        self.nav_bar = NavigationBar()
        self.nav_bar.bind(on_home=self.on_pre_enter)
        self.nav_bar.bind(on_add=self.go_to_add_subject)
        self.nav_bar.bind(on_menu=self.toggle_nav_drawer)

        self.nav_drawer = AppNavigationDrawer(parent_screen=self)
        main_content_area.add_widget(self.nav_drawer)

        self.layout.add_widget(main_content_area)
        self.layout.add_widget(self.nav_bar)
        self.add_widget(self.layout)

    def toggle_nav_drawer(self, *args):
        self.nav_drawer.set_state("open")

    def handle_swipe_left(self):
        self.go_to_add_subject()

    def update_datetime(self, dt):
        now = datetime.now()
        self.datetime_label.text = now.strftime("%A, %B %d, %Y | %I:%M:%S %p")

    def load_subject_cards(self):
        self.subject_box.clear_widgets()
        self.subjects = []
        professor_name = App.get_running_app().current_professor_name
        if not professor_name:
            return
        conn = sqlite3.connect("presko.db")
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT subject_code, subject_name, professor_name, schedule, section
            FROM subjects WHERE professor_name = ?
        """,
            (professor_name,),
        )
        self.subjects = cursor.fetchall()
        conn.close()
        if not self.subjects:
            # FIX: If the list is empty, fill the container and center the label.
            # 1. Make the layout fill the vertical space.
            self.subject_box.size_hint_y = 1
            # 2. Remove padding to allow for true centering.
            self.subject_box.padding = 0

            # 3. Add a top spacer widget to push the label down.
            self.subject_box.add_widget(Widget())
            # 4. Add the actual label.
            self.subject_box.add_widget(
                Label(
                    text="No subjects created yet.\nTap the '+' button to create one.",
                    color=(1, 1, 1, 1),
                    halign="center",
                )
            )
            # 5. Add a bottom spacer to push the label up.
            self.subject_box.add_widget(Widget())
        else:
            # If the list has items, ensure it's a normal scrollable list.
            # 1. Allow the layout to have a dynamic height for scrolling.
            self.subject_box.size_hint_y = None
            # 2. Set padding for the subject cards.
            self.subject_box.padding = (dp(10), dp(10))

            for data in self.subjects:
                text = (
                    f"[b]SUBJECT CODE:[/b] {data[0]}\n"
                    f"[b]SUBJECT NAME:[/b] {data[1]}\n"
                    f"[b]SCHEDULE:[/b] {data[3]}\n"
                    f"[b]SECTION:[/b] {data[4]}"
                )
                card = SubjectCard(
                    text=text,
                    size_hint_y=None,
                    height=dp(120),
                )
                card.bind(on_release=lambda x, code=data[0]: self.open_calendar(code))
                self.subject_box.add_widget(card)

    def open_calendar(self, subject_code):
        calendar_screen = self.manager.get_screen("calendar")
        calendar_screen.current_subject_code = subject_code
        self.manager.current = "calendar"

    def go_to_add_subject(self, *args):
        self.manager.transition.direction = "left"
        self.manager.current = "add_subject"

    def confirm_sign_out(self, *args):
        content = BoxLayout(orientation="vertical", spacing=10, padding=20)
        content.add_widget(
            Label(
                text="Are you sure you want to sign out?", color=(1, 1, 1, 1)
            )  # MODIFIED
        )
        btn_layout = BoxLayout(spacing=10, size_hint_y=None, height=dp(40))
        yes_btn = RoundedButton(
            text="Yes", bg_color=(0.8, 0.2, 0.2, 1), color=(1, 1, 1, 1)
        )
        no_btn = RoundedButton(text="No")
        btn_layout.add_widget(yes_btn)
        btn_layout.add_widget(no_btn)
        content.add_widget(btn_layout)
        self.confirm_popup = Popup(
            title="Confirm Sign Out",
            title_color=(1, 1, 1, 1),  # MODIFIED
            content=content,
            size_hint=(0.8, 0.3),
            background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
            separator_color=(0.5, 0, 0, 1),
        )
        yes_btn.bind(on_release=self.sign_out)
        no_btn.bind(on_release=self.confirm_popup.dismiss)
        self.confirm_popup.open()

    def sign_out(self, *args):
        self.confirm_popup.dismiss()
        self.manager.get_screen("login").set_user_type(None)
        self.manager.current = "welcome"

    def open_feedback_popup(self, role):
        layout = BoxLayout(orientation="vertical", spacing=10, padding=20)
        input_box = TextInput(
            hint_text="Write your feedback here...", multiline=True, size_hint=(1, 0.7)
        )
        layout.add_widget(input_box)
        status_label = Label(
            text="", size_hint=(1, 0.1), color=(1, 1, 1, 1)
        )  # MODIFIED
        layout.add_widget(status_label)
        submit_btn = RoundedButton(
            text="Submit Feedback",
            size_hint=(1, 0.2),
            bg_color=(0.1, 0.5, 0.1, 1),
            color=(1, 1, 1, 1),
        )
        layout.add_widget(submit_btn)
        popup = Popup(
            title="Send Feedback",
            title_color=(1, 1, 1, 1),  # MODIFIED
            content=layout,
            size_hint=(0.9, 0.6),
            background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
            separator_color=(0.5, 0, 0, 1),
        )

        def submit_feedback(instance):
            message = input_box.text.strip()
            if not message:
                status_label.text = "Feedback cannot be empty."
                return
            send_feedback(App.get_running_app().current_user_id, role, message)
            status_label.text = "Feedback sent successfully!"
            Clock.schedule_once(lambda dt: popup.dismiss(), 1.5)

        submit_btn.bind(on_release=submit_feedback)
        popup.open()

    def on_pre_enter(self, *args):
        self.nav_bar.update_active_item_visuals("Home")
        self.load_subject_cards()

    def on_leave(self, *args):
        """Closes the navigation drawer when the screen is left."""
        self.nav_drawer.set_state("close")


# ### MODIFIED ###: This screen now auto-generates a subject code.
class AddSubjectScreen(Screen, SwipeBehavior):
    user_role = StringProperty("Professor")

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.layout = BoxLayout(orientation="vertical")
        self.inputs = {}

        main_content_area = FloatLayout()
        main_content_area.add_widget(Image(source="Images/bg_app.jpg", fit_mode="fill"))
        header_bar = HeaderBar()
        main_content_area.add_widget(header_bar)

        main_content_area.add_widget(
            Image(
                source="Images/preskokooko-letter.png",
                size_hint=(None, None),
                size=(350, 350),
                pos_hint={"center_x": 0.5, "top": 1},
            )
        )
        self.datetime_label = Label(
            text="",
            size_hint=(None, None),
            size=(400, 40),
            pos_hint={"center_x": 0.5, "top": 0.97},
            font_size=18,
            color=(1, 1, 1, 1),  # MODIFIED: Changed from black to white
        )
        main_content_area.add_widget(self.datetime_label)
        Clock.schedule_interval(self.update_datetime, 1)

        form_container = BigRoundedCard(
            size=(dp(300), dp(280)),
            pos_hint={"center_x": 0.5, "center_y": 0.40},
            padding=dp(20),
            spacing=dp(10),
        )

        form_container.add_widget(
            Label(
                text="Create a New Subject",
                font_size="22sp",
                color=(0.5, 0, 0, 1),
                bold=True,
                size_hint_y=None,
                height=dp(30),
            )
        )

        # Inputs without the subject code field
        self.inputs = {
            "Subject Name": TextInput(hint_text="Insert Subject Name"),
            "Schedule": TextInput(hint_text="Time & Day Schedule (e.g., MW 9-10:30)"),
            "Section": TextInput(hint_text="Course, Year & Section (e.g., BSCS 3-1)"),
        }
        for field in self.inputs.values():
            field.size_hint_y = None
            field.height = dp(40)
            field.font_size = dp(14)
            form_container.add_widget(field)

        submit_btn = RoundedButton(
            text="Create Subject and Get Code",
            size_hint_y=None,
            height=dp(40),
            bg_color=(0.5, 0, 0, 1),
            color=(1, 1, 1, 1),
        )
        submit_btn.bind(on_release=self.submit)

        form_container.add_widget(submit_btn)

        main_content_area.add_widget(form_container)

        self.layout.add_widget(main_content_area)

        self.nav_bar = NavigationBar()
        self.nav_bar.bind(on_home=self.go_to_home)
        self.nav_bar.bind(on_add=lambda x: None)
        self.nav_bar.bind(on_menu=self.toggle_nav_drawer)

        self.nav_drawer = AppNavigationDrawer(parent_screen=self)
        main_content_area.add_widget(self.nav_drawer)

        self.layout.add_widget(self.nav_bar)
        self.add_widget(self.layout)

    def toggle_nav_drawer(self, *args):
        self.nav_drawer.set_state("open")

    def handle_swipe_right(self):
        self.go_to_home()

    def go_to_home(self, *args):
        self.manager.transition.direction = "right"
        self.manager.current = "professor_home"

    def on_pre_enter(self, *args):
        self.nav_bar.update_active_item_visuals("Add")

        for field in self.inputs.values():
            field.text = ""

    def update_datetime(self, dt):
        now = datetime.now()
        self.datetime_label.text = now.strftime("%A, %B %d, %Y | %I:%M:%S %p")

    def generate_unique_code(self, cursor):
        """Generates a 6-digit alphanumeric code that is unique in the subjects table."""
        while True:
            code = "".join(random.choices(string.ascii_uppercase + string.digits, k=6))
            cursor.execute("SELECT 1 FROM subjects WHERE subject_code = ?", (code,))
            if not cursor.fetchone():
                return code

    def open_feedback_popup(self, role):
        layout = BoxLayout(orientation="vertical", spacing=10, padding=20)
        input_box = TextInput(
            hint_text="Write your feedback here...", multiline=True, size_hint=(1, 0.7)
        )
        layout.add_widget(input_box)
        status_label = Label(
            text="", size_hint=(1, 0.1), color=(1, 1, 1, 1)
        )  # MODIFIED
        layout.add_widget(status_label)
        submit_btn = RoundedButton(
            text="Submit Feedback",
            size_hint=(1, 0.2),
            bg_color=(0.1, 0.5, 0.1, 1),
            color=(1, 1, 1, 1),
        )
        layout.add_widget(submit_btn)
        popup = Popup(
            title="Send Feedback",
            title_color=(1, 1, 1, 1),  # MODIFIED
            content=layout,
            size_hint=(0.9, 0.6),
            background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
            separator_color=(0.5, 0, 0, 1),
        )

        def submit_feedback(instance):
            message = input_box.text.strip()
            if not message:
                status_label.text = "Feedback cannot be empty."
                return
            send_feedback(App.get_running_app().current_user_id, role, message)
            status_label.text = "Feedback sent successfully!"
            Clock.schedule_once(lambda dt: popup.dismiss(), 1.5)

        submit_btn.bind(on_release=submit_feedback)
        popup.open()

    def confirm_sign_out(self, *args):
        content = BoxLayout(orientation="vertical", spacing=10, padding=20)
        content.add_widget(
            Label(
                text="Are you sure you want to sign out?", color=(1, 1, 1, 1)
            )  # MODIFIED
        )
        btn_layout = BoxLayout(spacing=10, size_hint_y=None, height=dp(40))
        yes_btn = RoundedButton(
            text="Yes", bg_color=(0.8, 0.2, 0.2, 1), color=(1, 1, 1, 1)
        )
        no_btn = RoundedButton(text="No")
        btn_layout.add_widget(yes_btn)
        btn_layout.add_widget(no_btn)
        content.add_widget(btn_layout)
        self.confirm_popup = Popup(
            title="Confirm Sign Out",
            title_color=(1, 1, 1, 1),  # MODIFIED
            content=content,
            size_hint=(0.8, 0.3),
            background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
            separator_color=(0.5, 0, 0, 1),
        )
        yes_btn.bind(on_release=self.sign_out)
        no_btn.bind(on_release=self.confirm_popup.dismiss)
        self.confirm_popup.open()

    def sign_out(self, *args):
        self.confirm_popup.dismiss()
        self.manager.get_screen("login").set_user_type(None)
        self.manager.current = "welcome"

    def submit(self, instance):
        subject_name = self.inputs["Subject Name"].text.strip()
        schedule = self.inputs["Schedule"].text.strip()
        section = self.inputs["Section"].text.strip().upper()
        professor_name = App.get_running_app().current_professor_name
        created_by = App.get_running_app().current_user_id

        if all([subject_name, schedule, section]):
            with sqlite3.connect("presko.db") as conn:
                cursor = conn.cursor()

                # Generate a unique code for the subject
                new_subject_code = self.generate_unique_code(cursor)

                cursor.execute(
                    """
                    INSERT INTO subjects (subject_code, subject_name, professor_name, schedule, section, created_by)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        new_subject_code,
                        subject_name,
                        professor_name,
                        schedule,
                        section,
                        created_by,
                    ),
                )

                # Show popup with the new code
                content = BoxLayout(
                    orientation="vertical", spacing=dp(10), padding=dp(10)
                )
                content.add_widget(
                    Label(
                        text="Subject Created!\nShare this code with your students:",
                        color=(1, 1, 1, 1),  # MODIFIED
                    )
                )
                content.add_widget(
                    Label(
                        text=f"[b]{new_subject_code}[/b]",
                        markup=True,
                        font_size="32sp",
                        color=(
                            0.9,
                            0.9,
                            0.9,
                            1,
                        ),  # MODIFIED: Brighter for dark background
                    )
                )
                ok_button = RoundedButton(text="OK", size_hint_y=None, height=dp(40))
                content.add_widget(ok_button)

                popup = Popup(
                    title="Subject Code",
                    content=content,
                    size_hint=(0.8, 0.4),
                    auto_dismiss=False,
                    background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
                    title_color=(1, 1, 1, 1),  # MODIFIED
                )

                def close_popup_and_go_home(instance):
                    popup.dismiss()
                    self.manager.current = "professor_home"

                ok_button.bind(on_release=close_popup_and_go_home)
                popup.open()

        else:
            popup = Popup(
                title="Missing Information",
                content=Label(
                    text="Please fill in all the fields.",
                    color=(1, 1, 1, 1),  # MODIFIED
                ),
                size_hint=(0.8, 0.3),
                background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
                title_color=(1, 1, 1, 1),  # MODIFIED
            )
            popup.open()

    def on_leave(self, *args):
        """Closes the navigation drawer when the screen is left."""
        self.nav_drawer.set_state("close")


# ==============================================================================
# =================== START: FULLY REPLACED/CORRECTED CLASS ====================
# ==============================================================================
class CalendarScreen(Screen, SwipeBehavior):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.layout = BoxLayout(orientation="vertical")
        self.current_subject_code = None
        self.current_year = datetime.now().year
        self.current_month = datetime.now().month
        self.selected_date = None

        main_content_area = FloatLayout()
        bg = Image(source="Images/bg_app.jpg", fit_mode="fill")
        main_content_area.add_widget(bg)
        header_bar = HeaderBar()
        main_content_area.add_widget(header_bar)

        main_content_area.add_widget(
            Image(
                source="Images/preskokooko-letter.png",
                size_hint=(None, None),
                size=(200, 200),
                pos_hint={"center_x": 0.5, "top": 1},
            )
        )
        self.datetime_label = Label(
            text="",
            size_hint=(None, None),
            size=(400, 40),
            pos_hint={"center_x": 0.5, "top": 0.97},
            font_size=18,
            color=(1, 1, 1, 1),  # MODIFIED: Changed from black to white
        )
        Clock.schedule_interval(self.update_datetime, 1)
        main_content_area.add_widget(self.datetime_label)

        rounded_container = BigRoundedCard(
            size=(dp(320), dp(450)),
            pos_hint={"center_x": 0.5, "center_y": 0.42},
            padding=[dp(20), dp(10), dp(20), dp(20)],
            spacing=dp(5),
        )
        nav_buttons = BoxLayout(size_hint_y=None, height=dp(40), spacing=5)
        self.month_label = Label(
            text=f"{calendar.month_name[self.current_month]} {self.current_year}",
            font_size=dp(16),
            color=(0.5, 0, 0, 1),
            bold=True,
        )
        prev_btn = Button(
            text="<",
            size_hint_x=0.2,
            background_color=(0.8, 0.8, 0.8, 1),
            color=(0, 0, 0, 1),
        )
        next_btn = Button(
            text=">",
            size_hint_x=0.2,
            background_color=(0.8, 0.8, 0.8, 1),
            color=(0, 0, 0, 1),
        )
        prev_btn.bind(on_release=self.go_to_prev_month)
        next_btn.bind(on_release=self.go_to_next_month)
        nav_buttons.add_widget(prev_btn)
        nav_buttons.add_widget(self.month_label)
        nav_buttons.add_widget(next_btn)
        rounded_container.add_widget(nav_buttons)

        self.calendar_grid = GridLayout(cols=7, spacing=dp(2))
        rounded_container.add_widget(self.calendar_grid)

        self.show_qr_btn = RoundedButton(
            text="Generate / Show QR Code",
            size_hint_y=None,
            height=dp(40),
            bg_color=(0.2, 0.5, 0.8, 1),
            color=(1, 1, 1, 1),
            disabled=True,
        )
        self.show_qr_btn.bind(on_release=self.show_qr_code)

        self.view_attendance_btn = RoundedButton(
            text="View Attendance List",
            size_hint_y=None,
            height=dp(40),
            bg_color=(0.1, 0.6, 0.2, 1),  # A green color
            color=(1, 1, 1, 1),
            disabled=True,
        )
        self.view_attendance_btn.bind(on_release=self.download_attendance_list)

        rounded_container.add_widget(Widget(size_hint_y=0.1))
        rounded_container.add_widget(self.show_qr_btn)
        rounded_container.add_widget(self.view_attendance_btn)
        main_content_area.add_widget(rounded_container)

        self.layout.add_widget(main_content_area)

        self.nav_bar = NavigationBar()
        self.nav_bar.bind(
            on_home=lambda x: setattr(self.manager, "current", "professor_home")
        )
        self.nav_bar.bind(
            on_add=lambda x: setattr(self.manager, "current", "add_subject")
        )

        self.layout.add_widget(self.nav_bar)
        self.add_widget(self.layout)

    def handle_swipe_right(self):
        self.manager.transition.direction = "right"
        self.manager.current = "professor_home"

    def on_pre_enter(self, *args):
        self.nav_bar.update_active_item_visuals("Home")

        load_qr_validity_from_db()
        today = datetime.now()
        self.current_year = today.year
        self.current_month = today.month
        self.month_label.text = (
            f"{calendar.month_name[self.current_month]} {self.current_year}"
        )
        if self.current_subject_code:
            self.generate_calendar(self.current_year, self.current_month)

    def update_datetime(self, dt):
        now = datetime.now()
        self.datetime_label.text = now.strftime("%A, %B %d, %Y | %I:%M:%S %p")

    def go_to_prev_month(self, instance):
        self.current_month -= 1
        if self.current_month == 0:
            self.current_month = 12
            self.current_year -= 1
        self.month_label.text = (
            f"{calendar.month_name[self.current_month]} {self.current_year}"
        )
        self.generate_calendar(self.current_year, self.current_month)

    def go_to_next_month(self, instance):
        self.current_month += 1
        if self.current_month == 13:
            self.current_month = 1
            self.current_year += 1
        self.month_label.text = (
            f"{calendar.month_name[self.current_month]} {self.current_year}"
        )
        self.generate_calendar(self.current_year, self.current_month)

    def generate_calendar(self, year, month):
        self.calendar_grid.clear_widgets()
        calendar.setfirstweekday(calendar.SUNDAY)
        for day in ["S", "M", "T", "W", "T", "F", "S"]:
            self.calendar_grid.add_widget(
                Label(text=day, color=(0.5, 0, 0, 1), bold=True, font_size=dp(12))
            )
        month_days = calendar.monthcalendar(year, month)
        marked_dates = marked_dates_per_subject.get(self.current_subject_code, set())
        for week in month_days:
            for day in week:
                if day == 0:
                    self.calendar_grid.add_widget(Widget())
                else:
                    date_str = f"{year}-{month:02d}-{day:02d}"
                    color = (
                        (0.9, 0.2, 0.2, 1)
                        if date_str in marked_dates
                        else (1, 1, 1, 0.8)
                    )
                    btn = Button(
                        text=str(day),
                        background_normal="",
                        background_color=color,
                        color=(0, 0, 0, 1),
                    )
                    btn.bind(on_release=partial(self.on_date_selected, date_str, btn))
                    self.calendar_grid.add_widget(btn)

        self.show_qr_btn.disabled = True
        self.view_attendance_btn.disabled = True

    def on_date_selected(self, date_str, button, *args):
        self.selected_date = date_str
        self.show_qr_btn.disabled = False

        key = f"{self.current_subject_code}_{self.selected_date}"
        if key in code_cache:
            self.view_attendance_btn.disabled = False
        else:
            self.view_attendance_btn.disabled = True

        if self.current_subject_code not in marked_dates_per_subject:
            marked_dates_per_subject[self.current_subject_code] = set()

        if date_str not in marked_dates_per_subject[self.current_subject_code]:
            marked_dates_per_subject[self.current_subject_code].add(date_str)
            button.background_color = (0.9, 0.2, 0.2, 1)

    def show_qr_code(self, instance):
        if not self.selected_date or not self.current_subject_code:
            return

        key = f"{self.current_subject_code}_{self.selected_date}"
        manual_code = code_cache.get(key)
        validity = qr_expiry_cache.get(key)

        is_currently_valid = False
        end_datetime_obj = None
        if validity:
            try:
                end_datetime_obj = datetime.strptime(
                    validity.get("end"), "%Y-%m-%d %H:%M:%S"
                )
                if datetime.now() < end_datetime_obj:
                    is_currently_valid = True
            except (ValueError, TypeError):
                is_currently_valid = False

        if is_currently_valid and manual_code:
            self._ask_regenerate_or_show(manual_code, end_datetime_obj)
        else:
            self._open_time_setting_popup()

    def _ask_regenerate_or_show(self, manual_code, end_datetime_obj):
        content = BoxLayout(orientation="vertical", padding=dp(20), spacing=dp(10))
        content.add_widget(
            Label(
                text="A valid QR code already exists for this day.\nWhat would you like to do?",
                color=(1, 1, 1, 1),  # MODIFIED
                halign="center",
            )
        )

        btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(45))
        show_btn = RoundedButton(text="Show Existing QR")
        new_btn = RoundedButton(
            text="Generate New QR", bg_color=(0.5, 0, 0, 1), color=(1, 1, 1, 1)
        )
        btn_layout.add_widget(show_btn)
        btn_layout.add_widget(new_btn)
        content.add_widget(btn_layout)

        popup = Popup(
            title="QR Code Found",
            content=content,
            size_hint=(0.9, 0.4),
            title_color=(1, 1, 1, 1),  # MODIFIED
            separator_color=(0.5, 0, 0, 1),
            background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
        )

        show_btn.bind(
            on_release=lambda x: (
                self._display_qr_popup(manual_code, end_datetime_obj),
                popup.dismiss(),
            )
        )
        new_btn.bind(
            on_release=lambda x: (self._open_time_setting_popup(), popup.dismiss())
        )
        popup.open()

    # ##################################################################################
    # ###################### START: MODIFIED AND CORRECTED METHOD ######################
    # ##################################################################################
    def _open_time_setting_popup(self):
        # The main content layout must be a FloatLayout to respect the pos_hint of its children.
        content = FloatLayout()

        # A vertical box to hold all the content, we will center this box.
        main_box = BoxLayout(
            orientation="vertical",
            padding=(dp(10), dp(20)),  # Adjusted padding
            spacing=dp(10),
            size_hint=(0.9, 0.9),  # Take up 90% of the popup space
            pos_hint={"center_x": 0.5, "center_y": 0.5},
        )

        # This is the container for our form fields.
        form_fields_box = BoxLayout(
            orientation="vertical",
            spacing=dp(10),
            size_hint_y=None,
            height=dp(135),  # Set a fixed height for the form area
        )

        now = datetime.now()

        # --- Row 1: Start Time ---
        row1 = BoxLayout(
            orientation="horizontal", spacing=dp(5), size_hint_y=None, height=dp(40)
        )
        row1.add_widget(
            Label(
                text="Start Time (HH:MM):",
                color=(1, 1, 1, 1),
                size_hint_x=0.6,
                halign="right",
                valign="middle",
            )
        )
        self.start_time_input = TextInput(
            text=now.strftime("%H:%M"),
            multiline=False,
            size_hint_x=0.4,  # Make the input box smaller
        )
        row1.add_widget(self.start_time_input)
        form_fields_box.add_widget(row1)

        # --- Row 2: Late After ---
        row2 = BoxLayout(
            orientation="horizontal", spacing=dp(5), size_hint_y=None, height=dp(40)
        )
        row2.add_widget(
            Label(
                text="Late after (minutes):",
                color=(1, 1, 1, 1),
                size_hint_x=0.6,
                halign="right",
                valign="middle",
            )
        )
        self.late_after_input = TextInput(
            text="15",
            multiline=False,
            input_filter="int",
            size_hint_x=0.4,  # Make the input box smaller
        )
        row2.add_widget(self.late_after_input)
        form_fields_box.add_widget(row2)

        # --- Row 3: Total Validity ---
        row3 = BoxLayout(
            orientation="horizontal", spacing=dp(5), size_hint_y=None, height=dp(40)
        )
        row3.add_widget(
            Label(
                text="Total validity (minutes):",
                color=(1, 1, 1, 1),
                size_hint_x=0.6,
                halign="right",
                valign="middle",
            )
        )
        self.total_validity_input = TextInput(
            text="60",
            multiline=False,
            input_filter="int",
            size_hint_x=0.4,  # Make the input box smaller
        )
        row3.add_widget(self.total_validity_input)
        form_fields_box.add_widget(row3)

        # Add the form fields to the main vertical box
        main_box.add_widget(form_fields_box)

        self.error_label = Label(
            text="",
            color=(1, 0.3, 0.3, 1),
            size_hint_y=None,
            height=dp(20),
        )
        main_box.add_widget(self.error_label)

        # Spacer to push the button down
        main_box.add_widget(Widget())

        generate_btn = RoundedButton(
            text="Generate QR",
            size_hint=(1, None),  # Stretch to fill the centered box
            height=dp(45),
            bg_color=(0.5, 0, 0, 1),
            color=(1, 1, 1, 1),
        )
        main_box.add_widget(generate_btn)

        # Add the centered box to the FloatLayout
        content.add_widget(main_box)

        self.time_popup = Popup(
            title="Set QR Code Validity",
            content=content,
            size_hint=(0.9, 0.6),  # Adjusted popup size
            title_color=(1, 1, 1, 1),
            separator_color=(0.5, 0, 0, 1),
            background_color=(0.2, 0.2, 0.2, 0.95),
        )

        generate_btn.bind(on_release=self._generate_qr_with_custom_time)
        self.time_popup.open()

    # ################################################################################
    # ####################### END: MODIFIED AND CORRECTED METHOD #####################
    # ################################################################################

    def _generate_qr_with_custom_time(self, instance):
        start_time_str = self.start_time_input.text.strip()
        late_after_str = self.late_after_input.text.strip()
        total_validity_str = self.total_validity_input.text.strip()

        if not start_time_str or not late_after_str or not total_validity_str:
            self.error_label.text = "All fields are required."
            return
        try:
            time_obj = time.strptime(start_time_str, "%H:%M")
            late_after_minutes = int(late_after_str)
            total_validity_minutes = int(total_validity_str)

            if late_after_minutes <= 0 or total_validity_minutes <= 0:
                self.error_label.text = "Time values must be positive."
                return
            if total_validity_minutes <= late_after_minutes:
                self.error_label.text = (
                    "Total validity must be greater than 'Late after'."
                )
                return

        except ValueError:
            self.error_label.text = "Invalid time format (HH:MM) or minutes."
            return

        self.time_popup.dismiss()

        selected_date_obj = datetime.strptime(self.selected_date, "%Y-%m-%d")

        start_datetime_obj = selected_date_obj.replace(
            hour=time_obj.tm_hour, minute=time_obj.tm_min
        )
        late_cutoff_datetime_obj = start_datetime_obj + timedelta(
            minutes=late_after_minutes
        )
        end_datetime_obj = start_datetime_obj + timedelta(
            minutes=total_validity_minutes
        )

        final_start_time_str = start_datetime_obj.strftime("%Y-%m-%d %H:%M:%S")
        final_end_time_str = end_datetime_obj.strftime("%Y-%m-%d %H:%M:%S")
        final_late_cutoff_time_str = late_cutoff_datetime_obj.strftime(
            "%Y-%m-%d %H:%M:%S"
        )

        key = f"{self.current_subject_code}_{self.selected_date}"
        manual_code = "".join(
            random.choices(string.ascii_uppercase + string.digits, k=6)
        )

        with sqlite3.connect("presko.db") as conn:
            cursor = conn.cursor()
            cursor.execute(
                "REPLACE INTO qr_validity (subject_code, date, manual_code, start_time, end_time, late_cutoff_time) VALUES (?, ?, ?, ?, ?, ?)",
                (
                    self.current_subject_code,
                    self.selected_date,
                    manual_code,
                    final_start_time_str,
                    final_end_time_str,
                    final_late_cutoff_time_str,
                ),
            )

        code_cache[key] = manual_code
        qr_expiry_cache[key] = {
            "start": final_start_time_str,
            "end": final_end_time_str,
            "late_cutoff": final_late_cutoff_time_str,
        }

        if self.current_subject_code not in marked_dates_per_subject:
            marked_dates_per_subject[self.current_subject_code] = set()
        marked_dates_per_subject[self.current_subject_code].add(self.selected_date)
        self.generate_calendar(self.current_year, self.current_month)
        self.view_attendance_btn.disabled = False

        self._display_qr_popup(manual_code, end_datetime_obj)

    def _display_qr_popup(self, manual_code, end_datetime_obj):
        qr = qrcode.make(manual_code)
        img_io = BytesIO()
        qr.save(img_io, format="PNG")
        img_io.seek(0)
        qr_texture = CoreImage(img_io, ext="png").texture

        layout = BoxLayout(orientation="vertical", padding=dp(20), spacing=dp(10))
        qr_image = Image(texture=qr_texture, size_hint_y=0.7)
        layout.add_widget(qr_image)
        layout.add_widget(
            Label(
                text=f"[b]Manual Code:[/b] {manual_code}",
                markup=True,
                font_size=dp(16),
                size_hint_y=0.1,
                color=(1, 1, 1, 1),  # MODIFIED
            )
        )
        layout.add_widget(
            Label(
                text=f"[b]Valid Until:[/b] {end_datetime_obj.strftime('%I:%M:%S %p')}",
                markup=True,
                font_size=dp(14),
                size_hint_y=0.1,
                color=(1, 1, 1, 1),  # MODIFIED
            )
        )
        popup = Popup(
            title="Scan for Attendance",
            title_color=(1, 1, 1, 1),  # MODIFIED
            content=layout,
            size_hint=(0.9, 0.6),
            background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
            separator_color=(0.5, 0, 0, 1),
        )
        popup.open()

    def download_attendance_list(self, instance):
        if not self.selected_date or not self.current_subject_code:
            return

        conn = sqlite3.connect("presko.db")
        cursor = conn.cursor()

        query = """
            SELECT
                u.username,
                u.name,
                a.status
            FROM users u
            INNER JOIN student_subjects ss ON u.username = ss.student_id
            LEFT JOIN attendance a ON u.username = a.student_id AND a.subject_code = ? AND a.date = ?
            WHERE ss.subject_code = ? AND u.role = 'Student'
            ORDER BY u.name;
        """
        cursor.execute(
            query,
            (self.current_subject_code, self.selected_date, self.current_subject_code),
        )
        records = cursor.fetchall()
        conn.close()

        if not records:
            popup = Popup(
                title="No Students",
                content=Label(
                    text="No students are enrolled in this subject.",
                    color=(1, 1, 1, 1),
                ),
                size_hint=(0.8, 0.3),
                background_color=(0.2, 0.2, 0.2, 0.95),
                title_color=(1, 1, 1, 1),
            )
            popup.open()
            return

        green_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        red_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        orange_fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        )
        green_font = Font(color="006100")
        red_font = Font(color="9C0006")
        orange_font = Font(color="9C5700")

        downloads_dir = "attendance_reports"
        if not os.path.exists(downloads_dir):
            os.makedirs(downloads_dir)

        filename = f"attendance_{self.current_subject_code}_{self.selected_date}.xlsx"
        filepath = os.path.join(downloads_dir, filename)

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance"
            header = ["Date", "ID Number", "Name", "Status"]
            ws.append(header)

            for cell in ws[1]:
                cell.font = Font(bold=True)

            for student_id, name, status in records:
                final_status = status if status else "absent"
                row_data = [
                    self.selected_date,
                    student_id,
                    name,
                    final_status.capitalize(),
                ]
                ws.append(row_data)
                last_row = ws.max_row
                status_cell = ws[f"D{last_row}"]

                if final_status.lower() == "present":
                    status_cell.fill = green_fill
                    status_cell.font = green_font
                elif final_status.lower() == "absent":
                    status_cell.fill = red_fill
                    status_cell.font = red_font
                elif final_status.lower() == "late":
                    status_cell.fill = orange_fill
                    status_cell.font = orange_font

            for col_idx, column_cells in enumerate(ws.columns, 1):
                padding = 5
                max_length = max(len(str(cell.value)) for cell in column_cells)
                adjusted_width = max_length + padding
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

            wb.save(filepath)

            # ###################### START: FIX ########################
            # Create a label that can wrap text for the popup message.
            popup_message = f"Attendance report saved successfully!\n\nLocation: {os.path.abspath(filepath)}"

            # 1. Create the Label widget instance
            popup_label = Label(
                text=popup_message,
                color=(1, 1, 1, 1),
                halign="center",
                valign="middle",  # Also good to vertically align
            )

            # 2. Bind the label's size to its text_size property.
            # This is the key to enabling text wrapping.
            popup_label.bind(size=popup_label.setter("text_size"))

            # 3. Use the new label as the popup's content.
            popup = Popup(
                title="Export Complete",
                content=popup_label,
                size_hint=(0.9, 0.5),  # Increased height slightly for better spacing
                background_color=(0.2, 0.2, 0.2, 0.95),
                title_color=(1, 1, 1, 1),
            )
            # ####################### END: FIX #########################

            popup.open()

        except Exception as e:
            # Create a label that can wrap text for the error popup
            error_message = f"Could not save file:\n{e}"
            error_label = Label(text=error_message, color=(1, 1, 1, 1))
            error_label.bind(size=error_label.setter("text_size"))

            popup = Popup(
                title="Export Error",
                content=error_label,
                size_hint=(0.9, 0.4),
                background_color=(0.2, 0.2, 0.2, 0.95),
                title_color=(1, 1, 1, 1),
            )
            popup.open()


# ==============================================================================
# =========================== END: REPLACED CLASS ==============================
# ==============================================================================


class ForgotPasswordScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.layout = FloatLayout()
        background = Image(source="Images/bg_app.jpg", fit_mode="fill")
        self.layout.add_widget(background)
        header_bar = HeaderBar()
        self.layout.add_widget(header_bar)
        self.layout.add_widget(
            Image(
                source="Images/preskokooko-letter.png",
                size_hint=(None, None),
                size=(380, 380),
                pos_hint={"center_x": 0.5, "top": 1.05},
            )
        )
        greeting = Label(
            text="[color=ffffff]Greetings, Iskolar ng Bayan![/color]",
            markup=True,
            font_size="22sp",
            pos_hint={"center_x": 0.5, "center_y": 0.75},
        )
        self.layout.add_widget(greeting)
        content_card = BigRoundedCard(
            size=(dp(300), dp(380)),
            pos_hint={"center_x": 0.5, "center_y": 0.42},
            padding=[dp(20), dp(0), dp(20), dp(0)],
            spacing=dp(15),
        )
        reset_password_label = Label(
            text="[b]Reset Password[/b]",
            markup=True,
            font_size="24sp",
            color=(0.5, 0, 0, 1),
            size_hint_y=None,
            height=dp(40),
        )
        content_card.add_widget(reset_password_label)
        self.username = TextInput(
            hint_text="PUP ID Number",
            multiline=False,
            size_hint_y=None,
            height=dp(40),
            padding=[10, 10],
        )
        self.new_password = TextInput(
            hint_text="New Password",
            password=True,
            multiline=False,
            size_hint_y=None,
            height=dp(40),
            padding=[10, 10],
        )
        self.confirm_password = TextInput(
            hint_text="Confirm New Password",
            password=True,
            multiline=False,
            size_hint_y=None,
            height=dp(40),
            padding=[10, 10],
        )
        self.role_spinner = OutlinedSpinner(
            text="Select Role",
            values=("Student", "Professor", "Admin"),
            size_hint_y=None,
            height=dp(40),
            font_size=16,
            option_cls=CustomSpinnerOption,
        )
        reset_btn = RoundedButton(
            text="RESET PASSWORD",
            size_hint_y=None,
            height=dp(45),
            bg_color=(0.5, 0, 0, 1),
            color=(1, 1, 1, 1),
            font_size=18,
        )
        reset_btn.bind(on_press=self.reset_password)
        self.msg = Label(
            size_hint_y=None, height=dp(30), color=(0.8, 0.2, 0.2, 1), bold=True
        )
        content_card.add_widget(self.username)
        content_card.add_widget(self.new_password)
        content_card.add_widget(self.confirm_password)
        content_card.add_widget(self.role_spinner)
        content_card.add_widget(reset_btn)
        content_card.add_widget(self.msg)
        self.layout.add_widget(content_card)
        back_btn = ImageButton(
            source="Images/back-button.png",
            size_hint=(None, None),
            size=(45, 45),
            pos_hint={"top": 0.98, "x": 0.01},
        )
        back_btn.bind(on_release=lambda x: setattr(self.manager, "current", "welcome"))
        self.layout.add_widget(back_btn)
        self.add_widget(self.layout)

    def reset_password(self, instance):
        uname = self.username.text.strip()
        new_pword = self.new_password.text.strip()
        confirm_pword = self.confirm_password.text.strip()
        role = self.role_spinner.text.strip()

        if not all([uname, new_pword, confirm_pword, role != "Select Role"]):
            self.msg.text = "All fields are required."
            return
        if new_pword != confirm_pword:
            self.msg.text = "Passwords do not match."
            return

        with sqlite3.connect("presko.db") as conn:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT * FROM users WHERE username = ? AND role = ?", (uname, role)
            )
            if cursor.fetchone():
                cursor.execute(
                    "UPDATE users SET password = ? WHERE username = ? AND role = ?",
                    (new_pword, uname, role),
                )
                self.msg.text = "Password reset successfully!"
                Clock.schedule_once(
                    lambda dt: setattr(self.manager, "current", "welcome"), 1.5
                )
            else:
                self.msg.text = "User not found or incorrect role."

    def on_pre_enter(self, *args):
        self.username.text = ""
        self.new_password.text = ""
        self.confirm_password.text = ""
        self.role_spinner.text = "Select Role"
        self.msg.text = ""

    def update_datetime(self, dt):
        pass


class AdminHomeScreen(Screen, SwipeBehavior):
    user_role = StringProperty("Admin")

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.layout = BoxLayout(orientation="vertical")
        main_content_area = FloatLayout()
        bg = Image(source="Images/bg_app.jpg", fit_mode="fill")
        main_content_area.add_widget(bg)
        header_bar = HeaderBar()
        main_content_area.add_widget(header_bar)
        main_content_area.add_widget(
            Image(
                source="Images/preskokooko-letter.png",
                size_hint=(None, None),
                size=(200, 200),
                pos_hint={"center_x": 0.5, "top": 1},
            )
        )
        admin_label = Label(
            text="[b]ADMIN[/b]",
            markup=True,
            font_size="18sp",
            size_hint=(0.4, 0.05),
            pos_hint={"center_x": 0.5, "top": 0.80},
            color=(1, 1, 1, 1),
        )
        with admin_label.canvas.before:
            Color(0.5, 0, 0, 1)
            self.admin_label_bg = Rectangle(size=admin_label.size, pos=admin_label.pos)
        admin_label.bind(
            pos=lambda i, v: setattr(self.admin_label_bg, "pos", v),
            size=lambda i, v: setattr(self.admin_label_bg, "size", v),
        )
        main_content_area.add_widget(admin_label)
        self.datetime_label = Label(
            text="",
            size_hint=(None, None),
            size=(400, 40),
            pos_hint={"center_x": 0.5, "top": 0.97},
            font_size=18,
            color=(1, 1, 1, 1),  # MODIFIED: Changed from black to white
        )
        Clock.schedule_interval(self.update_datetime, 1)
        main_content_area.add_widget(self.datetime_label)
        scrollview = ScrollView(
            size_hint=(0.9, 0.6),
            pos_hint={"center_x": 0.5, "top": 0.72},
        )
        self.feedback_box = BoxLayout(
            orientation="vertical", size_hint_y=None, padding=(10, 10), spacing=10
        )
        self.feedback_box.bind(minimum_height=self.feedback_box.setter("height"))
        scrollview.add_widget(self.feedback_box)
        main_content_area.add_widget(scrollview)

        self.nav_bar = NavigationBar()
        self.nav_bar.bind(on_home=self.on_pre_enter)
        self.nav_bar.bind(on_add=lambda x: None)
        self.nav_bar.bind(on_menu=self.toggle_nav_drawer)

        self.nav_drawer = AppNavigationDrawer(parent_screen=self)
        main_content_area.add_widget(self.nav_drawer)

        self.layout.add_widget(main_content_area)
        self.layout.add_widget(self.nav_bar)
        self.add_widget(self.layout)

    def toggle_nav_drawer(self, *args):
        self.nav_drawer.set_state("open")

    def update_datetime(self, dt):
        now = datetime.now()
        self.datetime_label.text = now.strftime("%A, %B %d, %Y | %I:%M:%S %p")

    def load_feedback_cards(self):
        self.feedback_box.clear_widgets()
        with sqlite3.connect("presko.db") as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT f.sender_id, f.sender_role, f.message, f.timestamp, u.name
                FROM feedback f
                JOIN users u ON f.sender_id = u.username AND f.sender_role = u.role
                ORDER BY f.timestamp DESC
                """
            )
            feedback_data = cursor.fetchall()

        for sender_id, sender_role, message, timestamp, sender_name in feedback_data:
            lines = [
                f"[b]From:[/b] {sender_name} ({sender_role})",
                f"[b]ID:[/b] {sender_id}",
                f"[b]Time:[/b] {timestamp}",
                f"[b]Message:[/b] {message}",
            ]
            card = AdminInfoCard(lines=lines, size_hint_y=None, height=dp(120))
            self.feedback_box.add_widget(card)

    def confirm_sign_out(self, *args):
        content = BoxLayout(orientation="vertical", spacing=10, padding=20)
        content.add_widget(
            Label(
                text="Are you sure you want to sign out?", color=(1, 1, 1, 1)
            )  # MODIFIED
        )
        btn_layout = BoxLayout(spacing=10, size_hint_y=None, height=dp(40))
        yes_btn = RoundedButton(
            text="Yes", bg_color=(0.8, 0.2, 0.2, 1), color=(1, 1, 1, 1)
        )
        no_btn = RoundedButton(text="No")
        btn_layout.add_widget(yes_btn)
        btn_layout.add_widget(no_btn)
        content.add_widget(btn_layout)
        self.confirm_popup = Popup(
            title="Confirm Sign Out",
            title_color=(1, 1, 1, 1),  # MODIFIED
            content=content,
            size_hint=(0.8, 0.3),
            background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
            separator_color=(0.5, 0, 0, 1),
        )
        yes_btn.bind(on_release=self.sign_out)
        no_btn.bind(on_release=self.confirm_popup.dismiss)
        self.confirm_popup.open()

    def sign_out(self, *args):
        self.confirm_popup.dismiss()
        self.manager.get_screen("login").set_user_type(None)
        self.manager.current = "welcome"

    def open_feedback_popup(self, role):
        layout = BoxLayout(orientation="vertical", spacing=10, padding=20)
        input_box = TextInput(
            hint_text="Write your feedback here...", multiline=True, size_hint=(1, 0.7)
        )
        layout.add_widget(input_box)
        status_label = Label(
            text="", size_hint=(1, 0.1), color=(1, 1, 1, 1)
        )  # MODIFIED
        layout.add_widget(status_label)
        submit_btn = RoundedButton(
            text="Submit Feedback",
            size_hint=(1, 0.2),
            bg_color=(0.1, 0.5, 0.1, 1),
            color=(1, 1, 1, 1),
        )
        layout.add_widget(submit_btn)
        popup = Popup(
            title="Send Feedback",
            title_color=(1, 1, 1, 1),  # MODIFIED
            content=layout,
            size_hint=(0.9, 0.6),
            background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
            separator_color=(0.5, 0, 0, 1),
        )

        def submit_feedback(instance):
            message = input_box.text.strip()
            if not message:
                status_label.text = "Feedback cannot be empty."
                return
            send_feedback(App.get_running_app().current_user_id, role, message)
            status_label.text = "Feedback sent successfully!"
            Clock.schedule_once(lambda dt: popup.dismiss(), 1.5)

        submit_btn.bind(on_release=submit_feedback)
        popup.open()

    def on_pre_enter(self, *args):
        self.nav_bar.update_active_item_visuals("Home")
        self.load_feedback_cards()

    def on_leave(self, *args):
        """Closes the navigation drawer when the screen is left."""
        self.nav_drawer.set_state("close")


# --- REVISED AND CORRECTED: Profile Screen with Animated Background ---
class ProfileScreen(Screen):
    particles = ListProperty([])

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        # Main layout with a dark background
        self.layout = FloatLayout()
        with self.layout.canvas.before:
            # 1. Dark background color
            Color(25 / 255, 25 / 255, 25 / 255, 1)
            self.bg = Rectangle(size=self.size, pos=self.pos)

            # 2. Setup for animated particles
            self.particle_canvas = self.canvas
            self.setup_particles()

        self.layout.bind(size=self._update_bg, pos=self._update_bg)

        # --- ALL UI WIDGET CREATION MOVED HERE, OUT OF _update_bg ---

        # Content layout to hold all widgets
        content = BoxLayout(
            orientation="vertical",
            padding=(dp(20), dp(80), dp(20), dp(20)),
            spacing=dp(15),
        )

        # --- Top Section: Profile Picture and Name ---
        top_section = BoxLayout(
            orientation="vertical",
            size_hint_y=None,
            height=dp(230),
            spacing=dp(10),
            pos_hint={"top": 1},
        )
        self.profile_pic = CircularImage(
            source="https://upload.wikimedia.org/wikipedia/commons/8/85/Elon_Musk_Royal_Society_%28crop1%29.jpg",
            size_hint_y=None,
            height=dp(120),
        )
        top_section.add_widget(self.profile_pic)
        self.name_label = Label(
            text="Full Name",
            font_size="28sp",
            bold=True,
            size_hint_y=None,
            height=dp(35),
        )
        top_section.add_widget(self.name_label)
        self.username_label = Label(
            text="@username",
            font_size="16sp",
            color=(0.7, 0.7, 0.7, 1),
            size_hint_y=None,
            height=dp(20),
        )
        top_section.add_widget(self.username_label)
        content.add_widget(top_section)

        # --- Bio Section ---
        bio_label = Label(
            text="Full-Stack Developer | Python, JavaScript, & Rust | Building the future, one line of code at a time. #OpenSource #Tech",
            font_size="15sp",
            color=(0.85, 0.85, 0.85, 1),
            size_hint_y=None,
            halign="center",
            valign="top",
        )
        bio_label.bind(
            width=lambda *x: bio_label.setter("text_size")(
                bio_label, (bio_label.width, None)
            )
        )
        bio_label.bind(texture_size=bio_label.setter("size"))
        content.add_widget(bio_label)

        # --- Edit Profile Button ---
        edit_button = RoundedButton(
            text="Edit Profile",
            size_hint_y=None,
            height=dp(48),
            bg_color=(34 / 255, 142 / 255, 222 / 255, 1),
            color=(1, 1, 1, 1),
            radius=[24],
        )
        edit_button.bind(on_press=self.edit_profile)
        content.add_widget(edit_button)

        # Spacer to push everything up
        content.add_widget(Widget())

        self.layout.add_widget(content)
        self.add_widget(self.layout)

        # Back Button to return to the previous screen
        back_btn = ImageButton(
            source="Images/back-button.png",
            size_hint=(None, None),
            size=(45, 45),
            pos_hint={"top": 0.98, "x": 0.01},
        )
        back_btn.bind(on_release=self.go_back)
        self.add_widget(back_btn)
        # --- END OF MOVED WIDGET CREATION ---

    def _update_bg(self, instance, value):
        # This method now ONLY updates the background rectangle
        self.bg.pos = instance.pos
        self.bg.size = instance.size

    def setup_particles(self):
        """Initializes the particles with random properties."""
        with self.particle_canvas:
            for _ in range(20):  # Number of particles
                color = Color(0.5, 0.5, 0.5, random.uniform(0.05, 0.2))
                size = random.uniform(dp(20), dp(80))
                pos = (random.uniform(0, self.width), random.uniform(0, self.height))
                velocity = (random.uniform(-0.3, 0.3), random.uniform(-0.3, 0.3))

                particle = Ellipse(pos=pos, size=(size, size))
                self.particles.append(
                    {"widget": particle, "velocity": velocity, "color": color}
                )

    def update_particles(self, dt):
        """Animates the particles on the canvas."""
        for p in self.particles:
            widget = p["widget"]
            vx, vy = p["velocity"]

            widget.pos = (widget.pos[0] + vx, widget.pos[1] + vy)

            if widget.pos[0] > self.width:
                widget.pos = (-widget.size[0], widget.pos[1])
            elif widget.pos[0] < -widget.size[0]:
                widget.pos = (self.width, widget.pos[1])

            if widget.pos[1] > self.height:
                widget.pos = (widget.pos[0], -widget.size[1])
            elif widget.pos[1] < -widget.size[1]:
                widget.pos = (widget.pos[0], self.height)

    def go_back(self, instance):
        app = App.get_running_app()
        role = app.current_user_role
        if role == "Student":
            self.manager.current = "student_home"
        elif role == "Professor":
            self.manager.current = "professor_home"
        elif role == "Admin":
            self.manager.current = "admin_home"
        else:
            self.manager.current = "welcome"

    def on_pre_enter(self, *args):
        app = App.get_running_app()
        self.name_label.text = app.current_user_name
        self.username_label.text = f"@{app.current_user_id}"

    def on_leave(self, *args):
        Clock.unschedule(self.update_particles)

    def on_enter(self, *args):
        Clock.schedule_interval(self.update_particles, 1 / 60.0)

    def edit_profile(self, instance):
        popup = Popup(
            title="Edit Profile",
            content=Label(
                text="This feature is coming soon!", color=(1, 1, 1, 1)
            ),  # MODIFIED
            size_hint=(0.8, 0.3),
            background_color=(0.2, 0.2, 0.2, 0.95),  # MODIFIED
            title_color=(1, 1, 1, 1),  # MODIFIED
        )
        popup.open()


class PreskoApp(MDApp):
    current_user_id = StringProperty("")
    current_user_name = StringProperty("")
    current_user_role = StringProperty("")
    current_student_id = StringProperty("")
    current_student_name = StringProperty("")
    current_professor_name = StringProperty("")

    def build(self):
        self.theme_cls.primary_palette = "Red"
        self.theme_cls.theme_style = "Light"

        sm = ScreenManager(transition=SlideTransition())
        sm.add_widget(WelcomeScreen(name="welcome"))
        sm.add_widget(SignUpScreen(name="signup"))
        sm.add_widget(LoginScreen(name="login"))
        sm.add_widget(StudentHomeScreen(name="student_home"))
        sm.add_widget(AddSubjectStudentScreen(name="add_subject_student"))
        sm.add_widget(StudentCalendarScreen(name="student_calendar"))
        sm.add_widget(ProfessorHomeScreen(name="professor_home"))
        sm.add_widget(AddSubjectScreen(name="add_subject"))
        sm.add_widget(CalendarScreen(name="calendar"))
        sm.add_widget(ForgotPasswordScreen(name="forgot_password"))
        sm.add_widget(AdminHomeScreen(name="admin_home"))
        sm.add_widget(ProfileScreen(name="profile"))
        return sm

    def on_start(self):
        init_database()


if __name__ == "__main__":
    PreskoApp().run()
